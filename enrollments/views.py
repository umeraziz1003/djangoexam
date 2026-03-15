from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db import transaction
import json
import csv
from datetime import datetime, date
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.cache import never_cache

from admission.models import Student
from academics.models import Department, Semester, Session, Batch
from courses.models import CourseOffering
from .forms import EnrollmentForm
from .models import Enrollment
from accounts.permissions import can


def _dept_restrict(request, qs, field_path):
    if request.user.is_department_scoped() and request.user.department_id:
        return qs.filter(**{field_path: request.user.department_id})
    return qs


def _require_enrollments_read(request):
    if not can(request.user, "ENROLLMENTS", "read"):
        return redirect("accounts:dashboard")
    return None


@login_required(login_url="accounts:login_page")
@never_cache
def enrollments_view(request):
    deny = _require_enrollments_read(request)
    if deny:
        return deny
    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "bulk_add_batch":
            if not can(request.user, "ENROLLMENTS", "create"):
                return redirect("enrollments:enrollments")
            department_id = request.POST.get("department_id") or ""
            batch_id = request.POST.get("batch_id") or ""
            semester_id = request.POST.get("semester_id") or ""
            offering_id = request.POST.get("offering_id") or ""
            exclude_ids = request.POST.getlist("exclude_student_ids")

            if request.user.is_department_scoped() and request.user.department_id:
                department_id = str(request.user.department_id)

            if not (department_id and batch_id and semester_id and offering_id):
                messages.error(request, "All fields are required for bulk enrollment.")
                return redirect("enrollments:enrollments")

            batch = Batch.objects.filter(id=batch_id).first()
            semester = Semester.objects.select_related("batch").filter(id=semester_id).first()
            offering = CourseOffering.objects.select_related("semester", "course").filter(id=offering_id).first()

            if not batch or not semester or not offering:
                messages.error(request, "Invalid batch, semester, or offering.")
                return redirect("enrollments:enrollments")

            if str(semester.batch_id) != str(batch.id):
                messages.error(request, "Selected semester does not belong to the selected batch.")
                return redirect("enrollments:enrollments")
            if str(offering.semester_id) != str(semester.id):
                messages.error(request, "Selected offering does not belong to the selected semester.")
                return redirect("enrollments:enrollments")
            if str(offering.semester.batch_id) != str(batch.id):
                messages.error(request, "Selected offering does not belong to the selected batch.")
                return redirect("enrollments:enrollments")
            if batch.department_id and str(batch.department_id) != str(department_id):
                messages.error(request, "Selected batch does not belong to the selected department.")
                return redirect("enrollments:enrollments")

            students_qs = Student.objects.filter(is_active=True, batch_id=batch.id)
            if exclude_ids:
                students_qs = students_qs.exclude(id__in=exclude_ids)

            created = 0
            skipped = 0
            for student in students_qs:
                _, was_created = Enrollment.objects.get_or_create(
                    student=student,
                    course_offering=offering,
                )
                if was_created:
                    created += 1
                else:
                    skipped += 1

            if created:
                messages.success(request, f"Enrolled {created} students. Skipped {skipped} (already enrolled).")
            else:
                messages.warning(request, f"No enrollments created. Skipped {skipped} (already enrolled).")
            return redirect("enrollments:enrollments")

    search = request.GET.get("search", "").strip()
    student_id = request.GET.get("student_id", "")
    offering_id = request.GET.get("offering_id", "")

    qs = Enrollment.objects.select_related(
        "student",
        "student__batch",
        "course_offering",
        "course_offering__course",
        "course_offering__course__department",
        "course_offering__semester",
        "course_offering__session",
    ).order_by("-date_enrolled")
    qs = _dept_restrict(request, qs, "course_offering__course__department_id")

    if student_id:
        qs = qs.filter(student_id=student_id)
    if offering_id:
        qs = qs.filter(course_offering_id=offering_id)
    if search:
        qs = qs.filter(
            Q(student__roll_no__icontains=search)
            | Q(student__full_name__icontains=search)
            | Q(course_offering__course__course_code__icontains=search)
            | Q(course_offering__course__course_title__icontains=search)
            | Q(course_offering__instructor_name__icontains=search)
            | Q(course_offering__session__name__icontains=search)
        )

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "enrollments": page_obj,
        "page_obj": page_obj,
        "students": _dept_restrict(request, Student.objects.filter(is_active=True).order_by("roll_no"), "batch__department_id"),
        "offerings": _dept_restrict(request, CourseOffering.objects.select_related(
            "course",
            "semester",
            "session",
        ).order_by("course__course_code"), "course__department_id"),
        "departments": _dept_restrict(request, Department.objects.filter(is_active=True).order_by("name"), "id"),
        "batches": _dept_restrict(request, Batch.objects.all().order_by("start_date"), "department_id"),
        "semesters": _dept_restrict(request, Semester.objects.select_related("batch").all().order_by("semester_number"), "batch__department_id"),
        "search": search,
        "student_id": student_id,
        "offering_id": offering_id,
    }
    return render(request, "enrollments/enrollments.html", context)


@login_required(login_url="accounts:login_page")
@never_cache
def create_enrollment(request):
    deny = _require_enrollments_read(request)
    if deny:
        return deny
    if request.method == "POST":
        if not can(request.user, "ENROLLMENTS", "create"):
            return redirect("enrollments:enrollments")
        form = EnrollmentForm(request.POST)
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["student"].queryset = form.fields["student"].queryset.filter(batch__department_id=request.user.department_id)
            form.fields["course_offering"].queryset = form.fields["course_offering"].queryset.filter(course__department_id=request.user.department_id)
        if form.is_valid():
            try:
                form.save()
                return redirect("enrollments:enrollments")
            except IntegrityError:
                form.add_error(None, "This student is already enrolled in the selected course offering.")
    else:
        form = EnrollmentForm()
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["student"].queryset = form.fields["student"].queryset.filter(batch__department_id=request.user.department_id)
            form.fields["course_offering"].queryset = form.fields["course_offering"].queryset.filter(course__department_id=request.user.department_id)

    return render(request, "enrollments/create_enrollment.html", {"form": form})


@login_required(login_url="accounts:login_page")
@never_cache
def edit_enrollment(request, pk):
    deny = _require_enrollments_read(request)
    if deny:
        return deny
    enrollment = get_object_or_404(Enrollment, pk=pk)
    if request.method == "POST":
        if not can(request.user, "ENROLLMENTS", "update"):
            return redirect("enrollments:enrollments")
        form = EnrollmentForm(request.POST, instance=enrollment)
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["student"].queryset = form.fields["student"].queryset.filter(batch__department_id=request.user.department_id)
            form.fields["course_offering"].queryset = form.fields["course_offering"].queryset.filter(course__department_id=request.user.department_id)
        if form.is_valid():
            try:
                form.save()
                return redirect("enrollments:enrollments")
            except IntegrityError:
                form.add_error(None, "This student is already enrolled in the selected course offering.")
    else:
        form = EnrollmentForm(instance=enrollment)
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["student"].queryset = form.fields["student"].queryset.filter(batch__department_id=request.user.department_id)
            form.fields["course_offering"].queryset = form.fields["course_offering"].queryset.filter(course__department_id=request.user.department_id)

    return render(request, "enrollments/edit_enrollment.html", {
        "form": form,
        "enrollment": enrollment,
    })


@login_required(login_url="accounts:login_page")
@never_cache
def delete_enrollment(request, pk):
    deny = _require_enrollments_read(request)
    if deny:
        return deny
    enrollment = get_object_or_404(Enrollment, pk=pk)
    if not can(request.user, "ENROLLMENTS", "delete"):
        return redirect("enrollments:enrollments")
    if request.method == "POST":
        enrollment.delete()
        return redirect("enrollments:enrollments")
    return render(request, "enrollments/delete_enrollment.html", {"enrollment": enrollment})


@login_required(login_url="accounts:login_page")
@never_cache
def semester_courses_view(request):
    deny = _require_enrollments_read(request)
    if deny:
        return deny
    department_id = request.GET.get("department_id", "")
    batch_id = request.GET.get("batch_id", "")
    semester_id = request.GET.get("semester_id", "")
    if request.user.is_department_scoped() and request.user.department_id:
        if not department_id:
            department_id = str(request.user.department_id)

    batches = Batch.objects.all()
    semesters = Semester.objects.select_related("batch").all()
    batches = _dept_restrict(request, batches, "department_id")
    semesters = _dept_restrict(request, semesters, "batch__department_id")
    if department_id:
        batches = batches.filter(department_id=department_id)
        semesters = semesters.filter(batch__department_id=department_id)
    if batch_id:
        semesters = semesters.filter(batch_id=batch_id)

    offerings = CourseOffering.objects.select_related(
        "course",
        "semester",
        "session",
        "course__department",
        "semester__batch",
    ).order_by("course__course_code")
    offerings = _dept_restrict(request, offerings, "course__department_id")
    if department_id:
        offerings = offerings.filter(course__department_id=department_id)
    if batch_id:
        offerings = offerings.filter(semester__batch_id=batch_id)
    if semester_id:
        offerings = offerings.filter(semester_id=semester_id)

    rows = []
    for o in offerings:
        rows.append({
            "offering": o,
            "students_count": o.enrollments.count(),
        })

    return render(request, "enrollments/semester_courses.html", {
        "departments": _dept_restrict(request, Department.objects.filter(is_active=True).order_by("name"), "id"),
        "batches": batches.order_by("start_date"),
        "semesters": semesters.order_by("semester_number"),
        "department_id": department_id,
        "batch_id": batch_id,
        "semester_id": semester_id,
        "rows": rows,
    })


@login_required(login_url="accounts:login_page")
@never_cache
def course_students_view(request):
    deny = _require_enrollments_read(request)
    if deny:
        return deny
    department_id = request.GET.get("department_id", "")
    batch_id = request.GET.get("batch_id", "")
    semester_id = request.GET.get("semester_id", "")
    session_id = request.GET.get("session_id", "")
    offering_id = request.GET.get("offering_id", "")
    if request.user.is_department_scoped() and request.user.department_id:
        if not department_id:
            department_id = str(request.user.department_id)

    session_obj = None
    session_batch_ids = []
    session_dept_ids = []
    if session_id:
        session_obj = Session.objects.prefetch_related("batches", "departments").filter(id=session_id).first()
        if session_obj:
            session_batch_ids = list(session_obj.batches.values_list("id", flat=True))
            session_dept_ids = list(session_obj.departments.values_list("id", flat=True))

    offerings = CourseOffering.objects.select_related(
        "course", "session", "semester", "semester__batch", "course__department"
    ).order_by("course__course_code")
    offerings = _dept_restrict(request, offerings, "course__department_id")
    if department_id:
        offerings = offerings.filter(course__department_id=department_id)
    if batch_id:
        offerings = offerings.filter(semester__batch_id=batch_id)
    if semester_id:
        offerings = offerings.filter(semester_id=semester_id)
    if session_id:
        offerings = offerings.filter(session_id=session_id)
    if session_batch_ids:
        offerings = offerings.filter(semester__batch_id__in=session_batch_ids)
    if session_dept_ids:
        offerings = offerings.filter(course__department_id__in=session_dept_ids)

    enrollments_qs = Enrollment.objects.select_related(
        "student",
        "student__batch",
        "course_offering",
        "course_offering__course",
        "course_offering__semester",
        "course_offering__session",
    ).order_by("student__roll_no")
    enrollments_qs = _dept_restrict(request, enrollments_qs, "course_offering__course__department_id")

    if department_id:
        enrollments_qs = enrollments_qs.filter(course_offering__course__department_id=department_id)
    if batch_id:
        enrollments_qs = enrollments_qs.filter(course_offering__semester__batch_id=batch_id)
    if semester_id:
        enrollments_qs = enrollments_qs.filter(course_offering__semester_id=semester_id)
    if session_id:
        enrollments_qs = enrollments_qs.filter(course_offering__session_id=session_id)
    if offering_id:
        enrollments_qs = enrollments_qs.filter(course_offering_id=offering_id)

    # show unique students (avoid duplicates across multiple offerings)
    seen = set()
    enrollments = []
    for e in enrollments_qs:
        if e.student_id in seen:
            continue
        seen.add(e.student_id)
        enrollments.append(e)

    departments_qs = _dept_restrict(request, Department.objects.filter(is_active=True).order_by("name"), "id")
    batches_qs = _dept_restrict(request, Batch.objects.all().order_by("start_date"), "department_id")
    semesters_qs = _dept_restrict(request, Semester.objects.select_related("batch").all().order_by("semester_number"), "batch__department_id")
    if session_obj:
        if session_batch_ids:
            batches_qs = batches_qs.filter(id__in=session_batch_ids)
            semesters_qs = semesters_qs.filter(batch_id__in=session_batch_ids)
        if session_dept_ids:
            departments_qs = departments_qs.filter(id__in=session_dept_ids)
        elif session_batch_ids:
            departments_qs = departments_qs.filter(batches__id__in=session_batch_ids).distinct()

    return render(request, "enrollments/course_students.html", {
        "departments": departments_qs,
        "batches": batches_qs,
        "semesters": semesters_qs,
        "sessions": Session.objects.order_by("-start_date"),
        "department_id": department_id,
        "batch_id": batch_id,
        "semester_id": semester_id,
        "session_id": session_id,
        "offerings": offerings,
        "offering_id": offering_id,
        "enrollments": enrollments,
    })


@login_required(login_url="accounts:login_page")
@never_cache
def student_courses_view(request):
    deny = _require_enrollments_read(request)
    if deny:
        return deny
    department_id = request.GET.get("department_id", "")
    batch_id = request.GET.get("batch_id", "")
    student_id = request.GET.get("student_id", "")
    if request.user.is_department_scoped() and request.user.department_id:
        if not department_id:
            department_id = str(request.user.department_id)
    students = Student.objects.filter(is_active=True).order_by("roll_no")
    students = _dept_restrict(request, students, "batch__department_id")
    if department_id:
        students = students.filter(batch__department_id=department_id)
    if batch_id:
        students = students.filter(batch_id=batch_id)
    enrollments_qs = Enrollment.objects.select_related(
        "course_offering",
        "course_offering__course",
        "course_offering__semester",
        "course_offering__session",
        "student",
        "student__batch",
    ).order_by("student__roll_no", "course_offering__course__course_code")
    enrollments_qs = _dept_restrict(request, enrollments_qs, "student__batch__department_id")

    if department_id:
        enrollments_qs = enrollments_qs.filter(student__batch__department_id=department_id)
    if batch_id:
        enrollments_qs = enrollments_qs.filter(student__batch_id=batch_id)
    if student_id:
        enrollments_qs = enrollments_qs.filter(student_id=student_id)

    return render(request, "enrollments/student_courses.html", {
        "departments": _dept_restrict(request, Department.objects.filter(is_active=True).order_by("name"), "id"),
        "batches": _dept_restrict(request, Batch.objects.all().order_by("start_date"), "department_id"),
        "department_id": department_id,
        "batch_id": batch_id,
        "students": students,
        "student_id": student_id,
        "enrollments": enrollments_qs,
    })


_ENR_HEADER_MAP = {
    "student": "student",
    "student_id": "student",
    "roll_no": "student",
    "course_offering": "course_offering",
    "offering_id": "course_offering",
}


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _ENR_HEADER_MAP.get(text, text)


def _resolve_student(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = Student.objects.filter(is_active=True)
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(batch__department_id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    return qs.filter(roll_no__iexact=raw).first()


def _resolve_offering(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = CourseOffering.objects.select_related("course", "session", "semester")
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(course__department_id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    for o in qs:
        if str(o).lower() == raw.lower():
            return o
    return None


def _validate_enrollment_rows(rows, request):
    cleaned = []
    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_header(k): v for k, v in row.items() }
        student = _resolve_student(normalized.get("student"), request)
        offering = _resolve_offering(normalized.get("course_offering"), request)
        if not student:
            errors.append("Student is invalid.")
        if not offering:
            errors.append("Course Offering is invalid.")
        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "student": str(student.id) if student else (normalized.get("student") or ""),
                "course_offering": str(offering.id) if offering else (normalized.get("course_offering") or ""),
            },
            "clean": {
                "student": student,
                "course_offering": offering,
            },
        })
    return cleaned


def _create_enrollment(clean):
    Enrollment.objects.create(
        student=clean["student"],
        course_offering=clean["course_offering"],
    )


@login_required(login_url="accounts:login_page")
@never_cache
def enrollments_template_download(request):
    if not can(request.user, "ENROLLMENTS", "create"):
        return redirect("enrollments:enrollments")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enrollments"
        ws.append(["Student Roll No", "Course Offering"])
        ws.append(["CS-2024-001", "CS-101 (Fall 2024)"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="enrollments_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="enrollments_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Student Roll No", "Course Offering"])
        writer.writerow(["CS-2024-001", "CS-101 (Fall 2024)"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def enrollments_bulk_preview(request):
    if not can(request.user, "ENROLLMENTS", "create"):
        return redirect("enrollments:enrollments")
    if request.method != "POST":
        return redirect("enrollments:enrollments")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("enrollments:enrollments")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("enrollments:enrollments")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("enrollments:enrollments")

    validated = _validate_enrollment_rows(rows, request)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "student", "label": "Student", "type": "text"},
        {"key": "course_offering", "label": "Course Offering", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "Enrollment Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("enrollments:enrollments_bulk_commit").url,
        "back_url": redirect("enrollments:enrollments").url,
        "hidden_fields": [],
        "extra_badge": "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def enrollments_bulk_commit(request):
    if not can(request.user, "ENROLLMENTS", "create"):
        return redirect("enrollments:enrollments")
    if request.method != "POST":
        return redirect("enrollments:enrollments")

    rows_json = request.POST.get("rows_json", "")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("enrollments:enrollments")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("enrollments:enrollments")

    validated = _validate_enrollment_rows(rows, request)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "student", "label": "Student", "type": "text"},
            {"key": "course_offering", "label": "Course Offering", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "Enrollment Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("enrollments:enrollments_bulk_commit").url,
            "back_url": redirect("enrollments:enrollments").url,
            "hidden_fields": [],
            "extra_badge": "",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_enrollment(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("enrollments:enrollments")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_enrollment(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("enrollments:enrollments")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("enrollments:enrollments")

    messages.error(request, "No valid rows to import.")
    return redirect("enrollments:enrollments")


@login_required(login_url="accounts:login_page")
def ajax_batches(request):
    if not (can(request.user, "ENROLLMENTS", "read") or can(request.user, "RESULTS", "read") or can(request.user, "EXAMS", "read")):
        return JsonResponse({"results": []}, status=403)
    department_id = request.GET.get("department_id")
    session_id = request.GET.get("session_id")
    qs = Batch.objects.all().order_by("start_date")
    qs = _dept_restrict(request, qs, "department_id")
    if department_id:
        qs = qs.filter(department_id=department_id)
    if session_id:
        session = Session.objects.filter(id=session_id).prefetch_related("batches").first()
        if session:
            qs = qs.filter(id__in=session.batches.values_list("id", flat=True))
    data = [{"id": b.id, "label": str(b)} for b in qs]
    return JsonResponse({"results": data})


@login_required(login_url="accounts:login_page")
def ajax_semesters(request):
    if not (can(request.user, "ENROLLMENTS", "read") or can(request.user, "RESULTS", "read") or can(request.user, "EXAMS", "read")):
        return JsonResponse({"results": []}, status=403)
    batch_id = request.GET.get("batch_id")
    session_id = request.GET.get("session_id")
    qs = Semester.objects.select_related("batch").all().order_by("semester_number")
    qs = _dept_restrict(request, qs, "batch__department_id")
    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    if session_id:
        session = Session.objects.filter(id=session_id).prefetch_related("batches").first()
        if session:
            qs = qs.filter(batch_id__in=session.batches.values_list("id", flat=True))
    data = [{"id": s.id, "label": str(s)} for s in qs]
    return JsonResponse({"results": data})


@login_required(login_url="accounts:login_page")
def ajax_students(request):
    if not (can(request.user, "ENROLLMENTS", "read") or can(request.user, "RESULTS", "read") or can(request.user, "EXAMS", "read")):
        return JsonResponse({"results": []}, status=403)
    department_id = request.GET.get("department_id")
    batch_id = request.GET.get("batch_id")
    session_id = request.GET.get("session_id")
    qs = Student.objects.filter(is_active=True).order_by("roll_no")
    qs = _dept_restrict(request, qs, "batch__department_id")
    if department_id:
        qs = qs.filter(batch__department_id=department_id)
    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    if session_id:
        session = Session.objects.filter(id=session_id).prefetch_related("batches").first()
        if session:
            qs = qs.filter(batch_id__in=session.batches.values_list("id", flat=True))
    data = [{"id": s.id, "label": f"{s.roll_no} — {s.full_name}"} for s in qs]
    return JsonResponse({"results": data})


@login_required(login_url="accounts:login_page")
def ajax_offerings(request):
    if not (can(request.user, "ENROLLMENTS", "read") or can(request.user, "RESULTS", "read") or can(request.user, "EXAMS", "read")):
        return JsonResponse({"results": []}, status=403)
    department_id = request.GET.get("department_id")
    batch_id = request.GET.get("batch_id")
    semester_id = request.GET.get("semester_id")
    session_id = request.GET.get("session_id")

    qs = CourseOffering.objects.select_related("course", "session", "semester").order_by("course__course_code")
    qs = _dept_restrict(request, qs, "course__department_id")
    if department_id:
        qs = qs.filter(course__department_id=department_id)
    if batch_id:
        qs = qs.filter(semester__batch_id=batch_id)
    if semester_id:
        qs = qs.filter(semester_id=semester_id)
    if session_id:
        qs = qs.filter(session_id=session_id)

    data = [{
        "id": o.id,
        "label": f"{o.course.course_code} — {o.course.course_title} ({o.session.name})",
    } for o in qs]
    return JsonResponse({"results": data})

