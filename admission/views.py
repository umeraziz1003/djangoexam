from django.contrib.auth import get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.views.decorators.cache import never_cache
from django.db import IntegrityError, transaction

from django.db.models import Q
from academics.models import Department, Batch
from accounts.permissions import can, ensure_default_groups
from django.contrib.auth.models import Group
from .forms import StudentForm
from .models import Student
import json
import csv
import re
from datetime import date, datetime

_STUDENT_HEADER_MAP = {
    "full_name": "full_name",
    "fullname": "full_name",
    "student_name": "full_name",
    "name": "full_name",
    "father_name": "father_name",
    "father": "father_name",
    "fathername": "father_name",
    "date_of_birth": "date_of_birth",
    "dob": "date_of_birth",
    "cnic": "cnic",
    "roll_no": "roll_no",
    "roll": "roll_no",
    "rollno": "roll_no",
    "registration_no": "registration_no",
    "registration": "registration_no",
    "reg_no": "registration_no",
    "regno": "registration_no",
    "batch": "batch",
    "batch_id": "batch",
    "batch_name": "batch",
}

_STUDENT_REQUIRED = [
    "full_name",
    "father_name",
    "date_of_birth",
    "cnic",
    "roll_no",
    "registration_no",
]


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", "_", text)
    return _STUDENT_HEADER_MAP.get(text, text)


def _parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _resolve_batch(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = Batch.objects.all()
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(department_id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    direct = qs.filter(name__iexact=raw).first() or qs.filter(title__iexact=raw).first()
    if direct:
        return direct
    raw_lower = raw.lower()
    for b in qs:
        if str(b).lower() == raw_lower:
            return b
    return None


def _validate_student_rows(rows, request, batch_override=None):
    cleaned = []
    cnics = []
    rolls = []
    for row in rows:
        cnic_val = (row.get("cnic") or "").strip()
        roll_val = (row.get("roll_no") or "").strip()
        if cnic_val:
            cnics.append(cnic_val)
        if roll_val:
            rolls.append(roll_val)

    existing_cnics = set(Student.objects.filter(cnic__in=cnics).values_list("cnic", flat=True))
    existing_rolls = set(Student.objects.filter(roll_no__in=rolls).values_list("roll_no", flat=True))
    User = get_user_model()
    existing_users = set(User.objects.filter(username__in=cnics).values_list("username", flat=True))

    seen_cnics = set()
    seen_rolls = set()

    for idx, row in enumerate(rows, start=1):
        row_errors = []
        normalized = {}
        for key, value in row.items():
            normalized[_normalize_header(key)] = value

        for field in _STUDENT_REQUIRED:
            if normalized.get(field) is None or str(normalized.get(field)).strip() == "":
                row_errors.append(f"{field.replace('_', ' ').title()} is required.")

        dob = _parse_date(normalized.get("date_of_birth"))
        if not dob:
            row_errors.append("Date of Birth is invalid or missing (use YYYY-MM-DD).")

        cnic_val = (normalized.get("cnic") or "").strip()
        roll_val = (normalized.get("roll_no") or "").strip()

        if cnic_val in seen_cnics:
            row_errors.append("CNIC is duplicated in the upload.")
        if roll_val in seen_rolls:
            row_errors.append("Roll No is duplicated in the upload.")

        if cnic_val in existing_cnics:
            row_errors.append("CNIC already exists.")
        if roll_val in existing_rolls:
            row_errors.append("Roll No already exists.")
        if cnic_val in existing_users:
            row_errors.append("User with this CNIC already exists.")

        batch_obj = batch_override or _resolve_batch(normalized.get("batch"), request)
        if not batch_obj:
            row_errors.append("Batch is invalid.")

        if cnic_val:
            seen_cnics.add(cnic_val)
        if roll_val:
            seen_rolls.add(roll_val)

        cleaned.append({
            "row_num": idx,
            "raw": normalized,
            "errors": row_errors,
            "clean": {
                "full_name": (normalized.get("full_name") or "").strip(),
                "father_name": (normalized.get("father_name") or "").strip(),
                "date_of_birth": dob,
                "cnic": cnic_val,
                "roll_no": roll_val,
                "registration_no": (normalized.get("registration_no") or "").strip(),
                "batch": batch_obj,
            },
        })

    return cleaned


def _create_student_from_clean(data):
    ensure_default_groups()
    student_group = Group.objects.filter(name="STUDENT").first()
    User = get_user_model()
    user = User.objects.create_user(
        username=data["cnic"],
        password=data["cnic"],
    )
    if student_group:
        user.groups.add(student_group)
    if data["batch"] and data["batch"].department_id:
        user.department_id = data["batch"].department_id
        user.save(update_fields=["department"])
    Student.objects.create(
        user=user,
        full_name=data["full_name"],
        father_name=data["father_name"],
        date_of_birth=data["date_of_birth"],
        cnic=data["cnic"],
        roll_no=data["roll_no"],
        registration_no=data["registration_no"],
        batch=data["batch"],
        is_active=True,
    )
# Create your views here.

@login_required(login_url="accounts:login_page")
@never_cache
def students_view(request):
    if not can(request.user, "STUDENTS", "read"):
        return redirect("accounts:dashboard")
    department_id = request.GET.get("department_id", "")
    batch_id = request.GET.get("batch_id", "")
    search = request.GET.get("search", "").strip()
    if request.user.is_department_scoped() and request.user.department_id:
        if not department_id:
            department_id = str(request.user.department_id)

    students = Student.objects.filter(is_active=True).select_related("batch", "batch__department")
    if department_id:
        students = students.filter(batch__department_id=department_id)
    if batch_id:
        students = students.filter(batch_id=batch_id)
    if search:
        students = students.filter(
            Q(full_name__icontains=search) | Q(roll_no__icontains=search)
        )

    departments = Department.objects.filter(is_active=True).order_by("name")
    batches = Batch.objects.all().order_by("start_date")
    if request.user.is_department_scoped() and request.user.department_id:
        departments = departments.filter(id=request.user.department_id)
        batches = batches.filter(department_id=request.user.department_id)
        students = students.filter(batch__department_id=request.user.department_id)

    return render(request, "admission/students.html", {
        "students": students,
        "departments": departments,
        "batches": batches,
        "department_id": department_id,
        "batch_id": batch_id,
        "search": search,
    })

@login_required(login_url="accounts:login_page")
@never_cache
def create_student(request):
    if not can(request.user, "STUDENTS", "create"):
        return redirect("admission:students")
    if request.method == "POST":
        form = StudentForm(request.POST)
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["batch"].queryset = Batch.objects.filter(department_id=request.user.department_id)
        if form.is_valid():
            cnic = form.cleaned_data.get("cnic")
            User = get_user_model()
            if User.objects.filter(username=cnic).exists():
                form.add_error("cnic", "A user with this CNIC already exists.")
            else:
                ensure_default_groups()
                student_group = Group.objects.filter(name="STUDENT").first()
                user = User.objects.create_user(
                    username=cnic,
                    password=cnic,
                )
                if student_group:
                    user.groups.add(student_group)
                student = form.save(commit=False)
                student.user = user
                if student.batch_id and student.batch.department_id:
                    user.department_id = student.batch.department_id
                    user.save(update_fields=["department"])
                student.save()
                return redirect("admission:students")
    else:
        form = StudentForm()
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["batch"].queryset = Batch.objects.filter(department_id=request.user.department_id)
    batches_qs = form.fields["batch"].queryset
    return render(request, "admission/create_student.html", {
        "form": form,
        "batches": batches_qs,
        "departments": Department.objects.filter(is_active=True).order_by("name"),
    })

@login_required(login_url="accounts:login_page")
@never_cache
def edit_student(request, pk):
    if not can(request.user, "STUDENTS", "update"):
        return redirect("admission:students")
    student = Student.objects.get(pk=pk)
    if request.method == "POST":
        form = StudentForm(request.POST, instance=student)
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["batch"].queryset = Batch.objects.filter(department_id=request.user.department_id)
        if form.is_valid():
            form.save()
            return redirect("admission:students")
    else:
        form = StudentForm(instance=student)
        if request.user.is_department_scoped() and request.user.department_id:
            form.fields["batch"].queryset = Batch.objects.filter(department_id=request.user.department_id)
    return render(request, "admission/edit_student.html", {"form": form, "student": student})

@login_required(login_url="accounts:login_page")
@never_cache
def delete_student(request, pk):
    if not can(request.user, "STUDENTS", "delete"):
        return redirect("admission:students")
    student = Student.objects.get(pk=pk)
    if request.method == "POST":
        student.is_active = False
        student.save()
        return redirect("admission:students")
    return render(request, "admission/delete_student.html", {"student": student})


@login_required(login_url="accounts:login_page")
@never_cache
def students_template_download(request):
    if not can(request.user, "STUDENTS", "create"):
        return redirect("admission:students")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append([
            "Full Name",
            "Father Name",
            "Date of Birth",
            "CNIC",
            "Roll No",
            "Registration No",
        ])
        ws.append([
            "Ali Raza",
            "Ahmed Raza",
            "2002-07-15",
            "35202-1234567-8",
            "CS-2024-001",
            "REG-2024-1001",
        ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="students_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="students_template.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Full Name",
            "Father Name",
            "Date of Birth",
            "CNIC",
            "Roll No",
            "Registration No",
        ])
        writer.writerow([
            "Ali Raza",
            "Ahmed Raza",
            "2002-07-15",
            "35202-1234567-8",
            "CS-2024-001",
            "REG-2024-1001",
        ])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def students_bulk_preview(request):
    if not can(request.user, "STUDENTS", "create"):
        return redirect("admission:students")
    if request.method != "POST":
        return redirect("admission:create_student")

    upload = request.FILES.get("excel_file")
    batch_override = None
    batch_id = (request.POST.get("batch_id") or "").strip()
    if not batch_id:
        messages.error(request, "Please select a batch before uploading.")
        return redirect("admission:create_student")
    batch_override = _resolve_batch(batch_id, request)
    if not batch_override:
        messages.error(request, "Selected batch is invalid.")
        return redirect("admission:create_student")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("admission:create_student")

    try:
        import openpyxl
    except ImportError:
        messages.error(request, "Excel import requires openpyxl. Please install it first.")
        return redirect("admission:create_student")

    try:
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("admission:create_student")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("admission:create_student")

    validated = _validate_student_rows(rows, request, batch_override=batch_override)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    return render(request, "admission/students_bulk_preview.html", {
        "rows": validated,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "batch_id": batch_id,
        "batch_display": str(batch_override) if batch_override else "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def students_bulk_commit(request):
    if not can(request.user, "STUDENTS", "create"):
        return redirect("admission:students")
    if request.method != "POST":
        return redirect("admission:create_student")

    rows_json = request.POST.get("rows_json", "")
    source = request.POST.get("source", "")
    batch_id = (request.POST.get("batch_id") or "").strip()
    if not batch_id:
        messages.error(request, "Please select a batch before importing.")
        return redirect("admission:create_student")
    batch_override = _resolve_batch(batch_id, request)
    if not batch_override:
        messages.error(request, "Selected batch is invalid.")
        return redirect("admission:create_student")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("admission:create_student")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("admission:create_student")

    validated = _validate_student_rows(rows, request, batch_override=batch_override)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        return render(request, "admission/students_bulk_preview.html", {
            "rows": validated,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": source,
            "batch_id": batch_id,
            "batch_display": str(batch_override) if batch_override else "",
        })

    created = 0
    failed = 0

    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_student_from_clean(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("admission:students")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_student_from_clean(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("admission:students")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("admission:create_student")

    messages.error(request, "No valid rows to import.")
    return redirect("admission:create_student")




