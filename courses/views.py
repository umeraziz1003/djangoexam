from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.cache import never_cache
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
import json
import csv
from datetime import datetime, date

from academics.models import Batch, Department, Semester, Session
from courses.models import Course
from courses.forms import CourseForm
from accounts.permissions import can
from courses.models import CourseOffering


def _require_courses_read(request):
    if not can(request.user, "COURSES", "read"):
        return redirect("accounts:dashboard")
    return None


def _require_offerings_read(request):
    if not can(request.user, "COURSE_OFFERINGS", "read"):
        return redirect("accounts:dashboard")
    return None


@login_required(login_url="accounts:login_page")
@never_cache
def courses_view(request):
    deny = _require_courses_read(request)
    if deny:
        return deny
    def _dept_restrict(qs, field_path):
        if request.user.is_department_scoped() and request.user.department_id:
            return qs.filter(**{field_path: request.user.department_id})
        return qs
    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "add_course":
            if not can(request.user, "COURSES", "create"):
                return redirect("courses:courses")
            course_code = request.POST.get("course_code", "").strip()
            course_title = request.POST.get("course_title", "").strip()
            credit_hours = request.POST.get("credit_hours", 3)
            course_type = request.POST.get("course_type", "Core")
            status = request.POST.get("status", "ACTIVE")
            department_id = request.POST.get("department_id")
            if request.user.is_department_scoped() and request.user.department_id:
                department_id = str(request.user.department_id)
            if course_code and course_title and department_id:
                Course.objects.create(
                    course_code=course_code,
                    course_title=course_title,
                    credit_hours=credit_hours,
                    course_type=course_type,
                    status=status,
                    department_id=department_id,
                )
            return redirect("courses:courses")

        elif action == "update_course":
            if not can(request.user, "COURSES", "update"):
                return redirect("courses:courses")
            course_id = request.POST.get("course_id")
            course = get_object_or_404(Course, pk=course_id)
            if request.user.is_department_scoped() and request.user.department_id:
                if course.department_id != request.user.department_id:
                    return redirect("courses:courses")
            course.course_code = request.POST.get("course_code", course.course_code).strip()
            course.course_title = request.POST.get("course_title", course.course_title).strip()
            try:
                course.credit_hours = int(request.POST.get("credit_hours", course.credit_hours))
            except (ValueError, TypeError):
                pass
            course.course_type = request.POST.get("course_type", course.course_type)
            course.status = request.POST.get("status", course.status)
            dept_id = request.POST.get("department_id")
            if dept_id and not (request.user.is_department_scoped() and request.user.department_id):
                course.department_id = dept_id
            course.save()
            return redirect("courses:courses")

        elif action == "delete_course":
            course_id = request.POST.get("course_id")
            course = get_object_or_404(Course, pk=course_id)
            if not can(request.user, "COURSES", "delete"):
                return redirect("courses:courses")
            if request.user.is_department_scoped() and request.user.department_id:
                if course.department_id != request.user.department_id:
                    return redirect("courses:courses")
            course.delete()
            return redirect("courses:courses")

        return redirect("courses:courses")

    # GET — apply filters
    department_id = request.GET.get("department_id", "")
    course_type = request.GET.get("course_type", "")
    search = request.GET.get("search", "").strip()

    if request.user.is_department_scoped() and request.user.department_id:
        if not department_id:
            department_id = str(request.user.department_id)
    qs = Course.objects.select_related("department").order_by("department__name", "course_code")
    qs = _dept_restrict(qs, "department_id")

    if department_id:
        qs = qs.filter(department_id=department_id)
    if course_type:
        qs = qs.filter(course_type=course_type)
    if search:
        qs = qs.filter(Q(course_code__icontains=search) | Q(course_title__icontains=search))

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    try:
        dept_id_int = int(department_id) if department_id else ""
    except (ValueError, TypeError):
        dept_id_int = ""

    context = {
        "courses": page_obj,
        "page_obj": page_obj,
        "departments": _dept_restrict(Department.objects.filter(is_active=True), "id"),
        "batches": Batch.objects.all(),
        "semesters": Semester.objects.all(),
        "department_id": dept_id_int,
        "course_type": course_type,
        "search": search,
    }
    return render(request, "courses/courses.html", context)


_COURSE_HEADER_MAP = {
    "department": "department",
    "department_id": "department",
    "department_code": "department",
    "course_code": "course_code",
    "code": "course_code",
    "course_title": "course_title",
    "title": "course_title",
    "credit_hours": "credit_hours",
    "credits": "credit_hours",
    "course_type": "course_type",
    "type": "course_type",
    "status": "status",
}


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _COURSE_HEADER_MAP.get(text, text)


def _resolve_department(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = Department.objects.filter(is_active=True)
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    return qs.filter(code__iexact=raw).first() or qs.filter(name__iexact=raw).first()


def _validate_course_rows(rows, request, department_override=None):
    cleaned = []
    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_header(k): v for k, v in row.items() }
        dept = department_override or _resolve_department(normalized.get("department"), request)
        course_code = (normalized.get("course_code") or "").strip()
        course_title = (normalized.get("course_title") or "").strip()
        course_type = (normalized.get("course_type") or "Core").strip()
        status = (normalized.get("status") or "ACTIVE").strip().upper()
        try:
            credit_hours = int(str(normalized.get("credit_hours")).strip())
        except (ValueError, TypeError):
            credit_hours = None

        if not dept:
            errors.append("Department is invalid.")
        if not course_code:
            errors.append("Course Code is required.")
        if not course_title:
            errors.append("Course Title is required.")
        if credit_hours is None:
            errors.append("Credit Hours is required.")
        if course_type not in {"Core", "Elective"}:
            course_type = "Core"
        if status not in {"ACTIVE", "INACTIVE"}:
            status = "ACTIVE"

        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "department": str(dept.id) if dept else (normalized.get("department") or ""),
                "course_code": course_code,
                "course_title": course_title,
                "credit_hours": "" if credit_hours is None else str(credit_hours),
                "course_type": course_type,
                "status": status,
            },
            "clean": {
                "department": dept,
                "course_code": course_code,
                "course_title": course_title,
                "credit_hours": credit_hours,
                "course_type": course_type,
                "status": status,
            },
        })
    return cleaned


def _create_course(clean):
    Course.objects.create(
        department=clean["department"],
        course_code=clean["course_code"],
        course_title=clean["course_title"],
        credit_hours=clean["credit_hours"],
        course_type=clean["course_type"],
        status=clean["status"],
    )


@login_required(login_url="accounts:login_page")
@never_cache
def courses_template_download(request):
    if not can(request.user, "COURSES", "create"):
        return redirect("courses:courses")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Courses"
        ws.append(["Course Code", "Course Title", "Credit Hours", "Course Type", "Status"])
        ws.append(["CS-101", "Programming Fundamentals", "3", "Core", "ACTIVE"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="courses_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="courses_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Course Code", "Course Title", "Credit Hours", "Course Type", "Status"])
        writer.writerow(["CS-101", "Programming Fundamentals", "3", "Core", "ACTIVE"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def courses_bulk_preview(request):
    if not can(request.user, "COURSES", "create"):
        return redirect("courses:courses")
    if request.method != "POST":
        return redirect("courses:courses")

    department_id = (request.POST.get("department_id") or "").strip()
    if not department_id:
        messages.error(request, "Please select a department before uploading.")
        return redirect("courses:add_course")
    dept_override = _resolve_department(department_id, request)
    if not dept_override:
        messages.error(request, "Selected department is invalid.")
        return redirect("courses:add_course")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("courses:add_course")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("courses:add_course")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("courses:add_course")

    validated = _validate_course_rows(rows, request, department_override=dept_override)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "course_code", "label": "Course Code", "type": "text"},
        {"key": "course_title", "label": "Course Title", "type": "text"},
        {"key": "credit_hours", "label": "Credit Hours", "type": "text"},
        {"key": "course_type", "label": "Course Type", "type": "text"},
        {"key": "status", "label": "Status", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "Course Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("courses:courses_bulk_commit").url,
        "back_url": redirect("courses:add_course").url,
        "hidden_fields": [{"name": "department_id", "value": department_id}],
        "extra_badge": f"Department: {dept_override}",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def courses_bulk_commit(request):
    if not can(request.user, "COURSES", "create"):
        return redirect("courses:courses")
    if request.method != "POST":
        return redirect("courses:courses")

    rows_json = request.POST.get("rows_json", "")
    department_id = (request.POST.get("department_id") or "").strip()
    if not department_id:
        messages.error(request, "Please select a department before importing.")
        return redirect("courses:add_course")
    dept_override = _resolve_department(department_id, request)
    if not dept_override:
        messages.error(request, "Selected department is invalid.")
        return redirect("courses:add_course")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("courses:add_course")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("courses:add_course")

    validated = _validate_course_rows(rows, request, department_override=dept_override)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "course_code", "label": "Course Code", "type": "text"},
            {"key": "course_title", "label": "Course Title", "type": "text"},
            {"key": "credit_hours", "label": "Credit Hours", "type": "text"},
            {"key": "course_type", "label": "Course Type", "type": "text"},
            {"key": "status", "label": "Status", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "Course Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("courses:courses_bulk_commit").url,
            "back_url": redirect("courses:add_course").url,
            "hidden_fields": [{"name": "department_id", "value": department_id}],
            "extra_badge": f"Department: {dept_override}",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_course(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("courses:courses")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_course(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("courses:courses")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("courses:courses")

    messages.error(request, "No valid rows to import.")
    return redirect("courses:courses")


@login_required(login_url="accounts:login_page")
@never_cache
def add_course_view(request):
    if not can(request.user, "COURSES", "create"):
        return redirect("courses:courses")
    def _dept_restrict(qs, field_path):
        if request.user.is_department_scoped() and request.user.department_id:
            return qs.filter(**{field_path: request.user.department_id})
        return qs
    if request.method == "POST":
        if not can(request.user, "COURSES", "create"):
            return redirect("courses:courses")
        data = request.POST.copy()
        if request.user.is_department_scoped() and request.user.department_id:
            data["department"] = request.user.department_id
        form = CourseForm(data)
        form.fields["department"].queryset = _dept_restrict(Department.objects.filter(is_active=True), "id")
        if form.is_valid():
            form.save()
            return redirect("courses:courses")
        return render(request, "courses/add_course.html", {
            "form": form,
            "departments": form.fields["department"].queryset,
        })

    form = CourseForm()
    form.fields["department"].queryset = _dept_restrict(Department.objects.filter(is_active=True), "id")
    if request.user.is_department_scoped() and request.user.department_id:
        form.fields["department"].initial = request.user.department_id
        form.fields["department"].disabled = True
    return render(request, "courses/add_course.html", {
        "form": form,
        "departments": form.fields["department"].queryset,
    })


@login_required(login_url="accounts:login_page")
@never_cache
def course_offerings_view(request):
    deny = _require_offerings_read(request)
    if deny:
        return deny
    from courses.models import CourseOffering

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "add_offering":
            if not can(request.user, "COURSE_OFFERINGS", "create"):
                return redirect("courses:course_offerings")
            course_id = request.POST.get("course_id")
            semester_id = request.POST.get("semester_id")
            session_id = request.POST.get("session_id")
            instructor_name = request.POST.get("instructor_name", "").strip()
            if course_id and semester_id and session_id and instructor_name:
                if request.user.is_department_scoped() and request.user.department_id:
                    course = get_object_or_404(Course, pk=course_id)
                    if course.department_id != request.user.department_id:
                        return redirect("courses:course_offerings")
                try:
                    CourseOffering.objects.create(
                        course_id=course_id,
                        semester_id=semester_id,
                        session_id=session_id,
                        instructor_name=instructor_name,
                    )
                except IntegrityError:
                    pass
            return redirect("courses:course_offerings")

        if action == "delete_offering":
            offering_id = request.POST.get("offering_id")
            if offering_id:
                offering = get_object_or_404(CourseOffering, pk=offering_id)
                if not can(request.user, "COURSE_OFFERINGS", "delete"):
                    return redirect("courses:course_offerings")
                if request.user.is_department_scoped() and request.user.department_id:
                    if offering.course.department_id != request.user.department_id:
                        return redirect("courses:course_offerings")
                offering.delete()
            return redirect("courses:course_offerings")

        if action == "bulk_add_offerings":
            if not can(request.user, "COURSE_OFFERINGS", "create"):
                return redirect("courses:course_offerings")
            department_id = request.POST.get("department_id") or ""
            batch_id = request.POST.get("batch_id") or ""
            semester_id = request.POST.get("semester_id") or ""
            session_id = request.POST.get("session_id") or ""
            course_ids = request.POST.getlist("course_ids")
            instructor_name = request.POST.get("instructor_name", "").strip()

            if not (department_id and batch_id and semester_id and session_id and course_ids and instructor_name):
                messages.error(request, "All fields are required for bulk add.")
                return redirect("courses:course_offerings")

            if request.user.is_department_scoped() and request.user.department_id:
                department_id = str(request.user.department_id)

            department = Department.objects.filter(id=department_id).first()
            semester = Semester.objects.select_related("batch").filter(id=semester_id).first()
            session_obj = Session.objects.filter(id=session_id).first()
            if not department or not semester or not session_obj:
                messages.error(request, "Invalid department, semester, or session.")
                return redirect("courses:course_offerings")

            if str(semester.batch_id) != str(batch_id):
                messages.error(request, "Selected semester does not belong to the selected batch.")
                return redirect("courses:course_offerings")
            if semester.batch.department_id != department.id:
                messages.error(request, "Selected batch does not belong to the selected department.")
                return redirect("courses:course_offerings")

            courses_qs = Course.objects.filter(id__in=course_ids, department_id=department.id)
            created = 0
            skipped = 0
            for course in courses_qs:
                try:
                    CourseOffering.objects.create(
                        course=course,
                        semester=semester,
                        session=session_obj,
                        instructor_name=instructor_name,
                    )
                    created += 1
                except IntegrityError:
                    skipped += 1
            if created:
                messages.success(request, f"Added {created} offerings. Skipped {skipped} (duplicates).")
            else:
                messages.warning(request, f"No offerings created. Skipped {skipped} (duplicates).")
            return redirect("courses:course_offerings")

        return redirect("courses:course_offerings")

    # GET — filters
    search = request.GET.get("search", "").strip()
    department_id = request.GET.get("department_id", "")
    batch_id = request.GET.get("batch_id", "")
    session_id = request.GET.get("session_id", "")
    semester_id = request.GET.get("semester_id", "")

    qs = CourseOffering.objects.select_related(
        "course",
        "course__department",
        "semester",
        "semester__batch",
        "session",
    ).order_by("-session__start_date", "course__course_code")
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(course__department_id=request.user.department_id)

    if session_id:
        qs = qs.filter(session_id=session_id)
    if semester_id:
        qs = qs.filter(semester_id=semester_id)
    if department_id:
        qs = qs.filter(course__department_id=department_id)
    if batch_id:
        qs = qs.filter(semester__batch_id=batch_id)
    if search:
        qs = qs.filter(
            Q(course__course_code__icontains=search)
            | Q(course__course_title__icontains=search)
            | Q(instructor_name__icontains=search)
        )

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    session_obj = None
    session_batch_ids = []
    session_dept_ids = []
    if session_id:
        session_obj = Session.objects.prefetch_related("batches", "departments").filter(id=session_id).first()
        if session_obj:
            session_batch_ids = list(session_obj.batches.values_list("id", flat=True))
            session_dept_ids = list(session_obj.departments.values_list("id", flat=True))

    context = {
        "offerings": page_obj,
        "page_obj": page_obj,
        "courses": Course.objects.select_related("department").order_by("course_code"),
        "semesters": Semester.objects.select_related("batch").all(),
        "sessions": Session.objects.order_by("-start_date"),
        "departments": Department.objects.filter(is_active=True),
        "batches": Batch.objects.all().order_by("start_date"),
        "search": search,
        "department_id": department_id,
        "batch_id": batch_id,
        "session_id": session_id,
        "semester_id": semester_id,
    }
    if session_obj:
        if session_batch_ids:
            context["batches"] = context["batches"].filter(id__in=session_batch_ids)
            context["semesters"] = context["semesters"].filter(batch_id__in=session_batch_ids)
        if session_dept_ids:
            context["departments"] = context["departments"].filter(id__in=session_dept_ids)
        elif session_batch_ids:
            context["departments"] = context["departments"].filter(batches__id__in=session_batch_ids).distinct()
    if request.user.is_department_scoped() and request.user.department_id:
        context["courses"] = context["courses"].filter(department_id=request.user.department_id)
        context["semesters"] = context["semesters"].filter(batch__department_id=request.user.department_id)
        context["departments"] = context["departments"].filter(id=request.user.department_id)
        context["batches"] = context["batches"].filter(department_id=request.user.department_id)
    return render(request, "courses/courses_offerings.html", context)


_OFFER_HEADER_MAP = {
    "course": "course",
    "course_id": "course",
    "course_code": "course",
    "semester": "semester",
    "semester_id": "semester",
    "session": "session",
    "session_id": "session",
    "instructor_name": "instructor_name",
    "instructor": "instructor_name",
}


def _normalize_offer_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _OFFER_HEADER_MAP.get(text, text)


def _resolve_course(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = Course.objects.all()
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(department_id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    return qs.filter(course_code__iexact=raw).first()


def _resolve_semester(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = Semester.objects.select_related("batch").all()
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(batch__department_id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    for s in qs:
        if str(s).lower() == raw.lower():
            return s
    return None


def _resolve_session(value):
    if value is None or str(value).strip() == "":
        return None
    raw = str(value).strip()
    if raw.isdigit():
        return Session.objects.filter(id=int(raw)).first()
    return Session.objects.filter(name__iexact=raw).first()


def _validate_offering_rows(rows, request):
    cleaned = []
    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_offer_header(k): v for k, v in row.items() }
        course = _resolve_course(normalized.get("course"), request)
        semester = _resolve_semester(normalized.get("semester"), request)
        session = _resolve_session(normalized.get("session"))
        instructor = (normalized.get("instructor_name") or "").strip()

        if not course:
            errors.append("Course is invalid.")
        if not semester:
            errors.append("Semester is invalid.")
        if not session:
            errors.append("Session is invalid.")
        if not instructor:
            errors.append("Instructor Name is required.")

        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "course": str(course.id) if course else (normalized.get("course") or ""),
                "semester": str(semester.id) if semester else (normalized.get("semester") or ""),
                "session": str(session.id) if session else (normalized.get("session") or ""),
                "instructor_name": instructor,
            },
            "clean": {
                "course": course,
                "semester": semester,
                "session": session,
                "instructor_name": instructor,
            },
        })
    return cleaned


def _create_offering(clean):
    CourseOffering.objects.create(
        course=clean["course"],
        semester=clean["semester"],
        session=clean["session"],
        instructor_name=clean["instructor_name"],
    )


@login_required(login_url="accounts:login_page")
@never_cache
def offerings_template_download(request):
    if not can(request.user, "COURSE_OFFERINGS", "create"):
        return redirect("courses:course_offerings")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Offerings"
        ws.append(["Course Code", "Semester", "Session", "Instructor Name"])
        ws.append(["CS-101", "BSCS 2022 - Semester 1", "Fall 2024", "Dr. Khan"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="offerings_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="offerings_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Course Code", "Semester", "Session", "Instructor Name"])
        writer.writerow(["CS-101", "BSCS 2022 - Semester 1", "Fall 2024", "Dr. Khan"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def offerings_bulk_preview(request):
    if not can(request.user, "COURSE_OFFERINGS", "create"):
        return redirect("courses:course_offerings")
    if request.method != "POST":
        return redirect("courses:course_offerings")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("courses:course_offerings")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("courses:course_offerings")
        header_keys = [_normalize_offer_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("courses:course_offerings")

    validated = _validate_offering_rows(rows, request)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "course", "label": "Course", "type": "text"},
        {"key": "semester", "label": "Semester", "type": "text"},
        {"key": "session", "label": "Session", "type": "text"},
        {"key": "instructor_name", "label": "Instructor Name", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "Offering Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("courses:offerings_bulk_commit").url,
        "back_url": redirect("courses:course_offerings").url,
        "hidden_fields": [],
        "extra_badge": "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def offerings_bulk_commit(request):
    if not can(request.user, "COURSE_OFFERINGS", "create"):
        return redirect("courses:course_offerings")
    if request.method != "POST":
        return redirect("courses:course_offerings")

    rows_json = request.POST.get("rows_json", "")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("courses:course_offerings")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("courses:course_offerings")

    validated = _validate_offering_rows(rows, request)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "course", "label": "Course", "type": "text"},
            {"key": "semester", "label": "Semester", "type": "text"},
            {"key": "session", "label": "Session", "type": "text"},
            {"key": "instructor_name", "label": "Instructor Name", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "Offering Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("courses:offerings_bulk_commit").url,
            "back_url": redirect("courses:course_offerings").url,
            "hidden_fields": [],
            "extra_badge": "",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_offering(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("courses:course_offerings")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_offering(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("courses:course_offerings")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("courses:course_offerings")

    messages.error(request, "No valid rows to import.")
    return redirect("courses:course_offerings")
