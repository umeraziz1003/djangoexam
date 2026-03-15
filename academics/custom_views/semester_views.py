from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.contrib import messages
from django.db import IntegrityError, transaction
import json
import csv
from datetime import datetime, date
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from ..forms import SemesterForm
from ..models import Batch, Department, Semester
from accounts.permissions import can


def _require_semesters_read(request):
    if not can(request.user, "SEMESTERS", "read"):
        return redirect("accounts:dashboard")
    return None


@login_required(login_url="accounts:login_page")
@never_cache
def semesters_view(request):
    deny = _require_semesters_read(request)
    if deny:
        return deny
    search = request.GET.get("search", "").strip()
    department_id = request.GET.get("department_id", "")
    batch_id = request.GET.get("batch_id", "")
    if request.user.is_department_scoped() and request.user.department_id:
        department_id = str(request.user.department_id)
    qs = Semester.objects.select_related("batch").order_by("batch", "semester_number")
    if department_id:
        qs = qs.filter(batch__department_id=department_id)
    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    if search:
        qs = qs.filter(
            Q(semester_number__icontains=search)
            | Q(batch__name__icontains=search)
            | Q(batch__title__icontains=search)
        )

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    form = SemesterForm()
    batches = Batch.objects.all().order_by("-start_date")
    if department_id:
        batches = batches.filter(department_id=department_id)
    departments = Department.objects.filter(is_active=True)
    if request.user.is_department_scoped() and request.user.department_id:
        departments = departments.filter(id=request.user.department_id)

    return render(request, "academics/semesters.html", {
        "semesters": page_obj,
        "page_obj": page_obj,
        "form": form,
        "search": search,
        "batches": batches,
        "departments": departments,
        "department_id": department_id,
        "batch_id": batch_id,
    })


_SEM_HEADER_MAP = {
    "batch": "batch",
    "batch_id": "batch",
    "batch_name": "batch",
    "semester_number": "semester_number",
    "semester": "semester_number",
    "semester_year": "semester_year",
    "year": "semester_year",
}


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _SEM_HEADER_MAP.get(text, text)


def _resolve_batch(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = Batch.objects.all()
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(department_id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    direct = qs.filter(name__iexact=raw).first() or qs.filter(title__iexact=raw).first()
    if direct:
        return direct
    raw_lower = raw.lower()
    for b in qs:
        if str(b).lower() == raw_lower:
            return b
    return None


def _validate_semester_rows(rows, request):
    cleaned = []
    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_header(k): v for k, v in row.items() }
        batch = _resolve_batch(normalized.get("batch"), request)
        sem_no = normalized.get("semester_number")
        sem_year = normalized.get("semester_year")
        try:
            sem_no_val = int(str(sem_no).strip())
        except (ValueError, TypeError):
            sem_no_val = None
        try:
            sem_year_val = int(str(sem_year).strip())
        except (ValueError, TypeError):
            sem_year_val = None

        if not batch:
            errors.append("Batch is invalid.")
        if sem_no_val is None:
            errors.append("Semester Number is required.")
        if sem_year_val is None:
            errors.append("Semester Year is required.")

        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "batch": str(batch.id) if batch else (normalized.get("batch") or ""),
                "semester_number": "" if sem_no_val is None else str(sem_no_val),
                "semester_year": "" if sem_year_val is None else str(sem_year_val),
            },
            "clean": {
                "batch": batch,
                "semester_number": sem_no_val,
                "semester_year": sem_year_val,
            },
        })
    return cleaned


def _create_semester(clean):
    Semester.objects.create(
        batch=clean["batch"],
        semester_number=clean["semester_number"],
        semester_year=clean["semester_year"],
    )


@login_required(login_url="accounts:login_page")
@never_cache
def semesters_template_download(request):
    deny = _require_semesters_read(request)
    if deny:
        return deny
    if not can(request.user, "SEMESTERS", "create"):
        return redirect("academics:semesters")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Semesters"
        ws.append(["Batch", "Semester Number", "Semester Year"])
        ws.append(["BSCS 2022", "1", "1"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="semesters_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="semesters_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Batch", "Semester Number", "Semester Year"])
        writer.writerow(["BSCS 2022", "1", "1"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def semesters_bulk_preview(request):
    deny = _require_semesters_read(request)
    if deny:
        return deny
    if not can(request.user, "SEMESTERS", "create"):
        return redirect("academics:semesters")
    if request.method != "POST":
        return redirect("academics:semesters")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("academics:semesters")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("academics:semesters")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("academics:semesters")

    validated = _validate_semester_rows(rows, request)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "batch", "label": "Batch", "type": "text"},
        {"key": "semester_number", "label": "Semester Number", "type": "text"},
        {"key": "semester_year", "label": "Semester Year", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "Semester Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("academics:semesters_bulk_commit").url,
        "back_url": redirect("academics:semesters").url,
        "hidden_fields": [],
        "extra_badge": "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def semesters_bulk_commit(request):
    deny = _require_semesters_read(request)
    if deny:
        return deny
    if not can(request.user, "SEMESTERS", "create"):
        return redirect("academics:semesters")
    if request.method != "POST":
        return redirect("academics:semesters")

    rows_json = request.POST.get("rows_json", "")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("academics:semesters")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("academics:semesters")

    validated = _validate_semester_rows(rows, request)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "batch", "label": "Batch", "type": "text"},
            {"key": "semester_number", "label": "Semester Number", "type": "text"},
            {"key": "semester_year", "label": "Semester Year", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "Semester Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("academics:semesters_bulk_commit").url,
            "back_url": redirect("academics:semesters").url,
            "hidden_fields": [],
            "extra_badge": "",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_semester(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("academics:semesters")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_semester(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("academics:semesters")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("academics:semesters")

    messages.error(request, "No valid rows to import.")
    return redirect("academics:semesters")


@login_required(login_url="accounts:login_page")
def create_semester(request):
    deny = _require_semesters_read(request)
    if deny:
        return deny
    if request.method == "POST":
        if not can(request.user, "SEMESTERS", "create"):
            return redirect("academics:semesters")
        form = SemesterForm(request.POST)
        if form.is_valid():
            if request.user.is_department_scoped() and request.user.department_id:
                if form.cleaned_data["batch"].department_id != request.user.department_id:
                    return redirect("academics:semesters")
            form.save()
            return redirect("academics:semesters")
        qs = Semester.objects.select_related("batch").order_by("batch", "semester_number")
        paginator = Paginator(qs, 10)
        page_obj = paginator.get_page(None)
        batches = Batch.objects.all().order_by("-start_date")
        return render(request, "academics/semesters.html", {
            "semesters": page_obj,
            "page_obj": page_obj,
            "form": form,
            "search": "",
            "batches": batches,
        })
    return redirect("academics:semesters")


@login_required(login_url="accounts:login_page")
def edit_semester(request, pk):
    deny = _require_semesters_read(request)
    if deny:
        return deny
    semester = get_object_or_404(Semester, pk=pk)
    if request.method == "POST":
        if not can(request.user, "SEMESTERS", "update"):
            return redirect("academics:semesters")
        if request.user.is_department_scoped() and request.user.department_id:
            if semester.batch.department_id != request.user.department_id:
                return redirect("academics:semesters")
        batch_id = request.POST.get("batch_id")
        if batch_id and not (request.user.is_department_scoped() and request.user.department_id):
            semester.batch_id = batch_id
        try:
            semester.semester_number = int(request.POST.get("semester_number", semester.semester_number))
            semester.semester_year = int(request.POST.get("semester_year", semester.semester_year))
        except (ValueError, TypeError):
            pass
        semester.save()
    return redirect("academics:semesters")


@login_required(login_url="accounts:login_page")
@require_POST
def delete_semester(request, pk):
    deny = _require_semesters_read(request)
    if deny:
        return deny
    semester = get_object_or_404(Semester, pk=pk)
    if not can(request.user, "SEMESTERS", "delete"):
        return redirect("academics:semesters")
    if request.user.is_department_scoped() and request.user.department_id:
        if semester.batch.department_id != request.user.department_id:
            return redirect("academics:semesters")
    semester.delete()
    return redirect("academics:semesters")
