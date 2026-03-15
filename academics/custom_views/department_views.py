from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.contrib import messages
from django.db import IntegrityError, transaction
import json
import csv
from datetime import datetime, date
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from ..forms import DepartmentForm
from ..models import Department
from accounts.permissions import can


@login_required(login_url="accounts:login_page")
@never_cache
def departments_view(request):
    if not can(request.user, "DEPARTMENTS", "read"):
        return redirect("accounts:dashboard")
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    search = request.GET.get("search", "").strip()
    qs = Department.objects.all().order_by("name")
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(code__icontains=search))

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    form = DepartmentForm()

    return render(request, "academics/departments.html", {
        "departments": page_obj,
        "page_obj": page_obj,
        "form": form,
        "search": search,
    })


_DEPT_HEADER_MAP = {
    "name": "name",
    "code": "code",
    "duration_years": "duration_years",
    "duration": "duration_years",
    "years": "duration_years",
    "is_active": "is_active",
    "active": "is_active",
}


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _DEPT_HEADER_MAP.get(text, text)


def _parse_bool(val):
    if val is None or val == "":
        return True
    raw = str(val).strip().lower()
    return raw in {"1", "true", "yes", "y", "active"}


def _validate_department_rows(rows):
    cleaned = []
    codes = []
    for r in rows:
        code = (r.get("code") or "").strip()
        if code:
            codes.append(code)
    existing_codes = set(Department.objects.filter(code__in=codes).values_list("code", flat=True))
    seen_codes = set()

    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_header(k): v for k, v in row.items() }
        name = (normalized.get("name") or "").strip()
        code = (normalized.get("code") or "").strip()
        duration = normalized.get("duration_years")
        if not name:
            errors.append("Name is required.")
        if not code:
            errors.append("Code is required.")
        if code in seen_codes:
            errors.append("Code is duplicated in the upload.")
        if code and code in existing_codes:
            errors.append("Code already exists.")
        try:
            duration_val = int(str(duration).strip())
        except (ValueError, TypeError):
            duration_val = None
        if duration_val is None:
            errors.append("Duration Years must be a number.")

        is_active = _parse_bool(normalized.get("is_active"))
        if code:
            seen_codes.add(code)

        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "name": name,
                "code": code,
                "duration_years": "" if duration_val is None else str(duration_val),
                "is_active": "true" if is_active else "false",
            },
            "clean": {
                "name": name,
                "code": code,
                "duration_years": duration_val,
                "is_active": is_active,
            },
        })
    return cleaned


def _create_department(clean):
    Department.objects.create(
        name=clean["name"],
        code=clean["code"],
        duration_years=clean["duration_years"],
        is_active=clean["is_active"],
    )


@login_required(login_url="accounts:login_page")
@never_cache
def departments_template_download(request):
    if not can(request.user, "DEPARTMENTS", "create"):
        return redirect("academics:departments")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Departments"
        ws.append(["Name", "Code", "Duration Years", "Is Active"])
        ws.append(["Computer Science", "CS", "4", "true"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="departments_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="departments_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Name", "Code", "Duration Years", "Is Active"])
        writer.writerow(["Computer Science", "CS", "4", "true"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def departments_bulk_preview(request):
    if not can(request.user, "DEPARTMENTS", "create"):
        return redirect("academics:departments")
    if request.method != "POST":
        return redirect("academics:departments")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("academics:departments")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("academics:departments")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("academics:departments")

    validated = _validate_department_rows(rows)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "name", "label": "Name", "type": "text"},
        {"key": "code", "label": "Code", "type": "text"},
        {"key": "duration_years", "label": "Duration Years", "type": "text"},
        {"key": "is_active", "label": "Is Active", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "Department Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("academics:departments_bulk_commit").url,
        "back_url": redirect("academics:departments").url,
        "hidden_fields": [],
        "extra_badge": "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def departments_bulk_commit(request):
    if not can(request.user, "DEPARTMENTS", "create"):
        return redirect("academics:departments")
    if request.method != "POST":
        return redirect("academics:departments")

    rows_json = request.POST.get("rows_json", "")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("academics:departments")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("academics:departments")

    validated = _validate_department_rows(rows)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "name", "label": "Name", "type": "text"},
            {"key": "code", "label": "Code", "type": "text"},
            {"key": "duration_years", "label": "Duration Years", "type": "text"},
            {"key": "is_active", "label": "Is Active", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "Department Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("academics:departments_bulk_commit").url,
            "back_url": redirect("academics:departments").url,
            "hidden_fields": [],
            "extra_badge": "",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_department(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("academics:departments")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_department(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("academics:departments")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("academics:departments")

    messages.error(request, "No valid rows to import.")
    return redirect("academics:departments")


@login_required(login_url="accounts:login_page")
def create_department(request):
    if request.method == "POST":
        if not request.user.is_exam_officer():
            return redirect("accounts:dashboard")
        if not can(request.user, "DEPARTMENTS", "create"):
            return redirect("academics:departments")
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("academics:departments")
        # Re-render with errors
        qs = Department.objects.all().order_by("name")
        paginator = Paginator(qs, 10)
        page_obj = paginator.get_page(None)
        return render(request, "academics/departments.html", {
            "departments": page_obj,
            "page_obj": page_obj,
            "form": form,
            "search": "",
        })
    return redirect("academics:departments")


@login_required(login_url="accounts:login_page")
def edit_department(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == "POST":
        if not request.user.is_exam_officer():
            return redirect("accounts:dashboard")
        if not can(request.user, "DEPARTMENTS", "update"):
            return redirect("academics:departments")
        department.name = request.POST.get("name", department.name).strip()
        department.code = request.POST.get("code", department.code).strip()
        try:
            department.duration_years = int(request.POST.get("duration_years", department.duration_years))
        except (ValueError, TypeError):
            pass
        department.is_active = "is_active" in request.POST
        department.save()
    return redirect("academics:departments")


@login_required(login_url="accounts:login_page")
@require_POST
def delete_department(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    if not can(request.user, "DEPARTMENTS", "delete"):
        return redirect("academics:departments")
    department.delete()
    return redirect("academics:departments")
