from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.contrib import messages
from django.db import IntegrityError, transaction
import json
import csv
from datetime import datetime, date
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from ..forms import SessionForm
from ..models import Session, Department, Batch
from accounts.permissions import can


def _require_sessions_read(request):
    if not can(request.user, "SESSIONS", "read"):
        return redirect("accounts:dashboard")
    return None


@login_required(login_url="accounts:login_page")
@never_cache
def sessions_view(request):
    deny = _require_sessions_read(request)
    if deny:
        return deny
    search = request.GET.get("search", "").strip()
    status = request.GET.get("status", "")
    qs = Session.objects.all().order_by("-start_date")
    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)
    if search:
        qs = qs.filter(Q(name__icontains=search))

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    form = SessionForm()

    return render(request, "academics/sessions.html", {
        "sessions": page_obj,
        "page_obj": page_obj,
        "form": form,
        "departments": Department.objects.filter(is_active=True).order_by("name"),
        "batches": Batch.objects.select_related("department").all().order_by("start_date"),
        "search": search,
        "status": status,
    })


_SESSION_HEADER_MAP = {
    "name": "name",
    "start_date": "start_date",
    "end_date": "end_date",
    "is_active": "is_active",
    "departments": "departments",
    "batches": "batches",
}


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _SESSION_HEADER_MAP.get(text, text)


def _parse_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    raw = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(val):
    if val is None or val == "":
        return True
    raw = str(val).strip().lower()
    return raw in {"1", "true", "yes", "y", "active"}


def _resolve_departments(value, request):
    if value is None or str(value).strip() == "":
        return []
    qs = Department.objects.filter(is_active=True)
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(id=request.user.department_id)
    items = [v.strip() for v in str(value).split(",") if v.strip()]
    result = []
    for item in items:
        dept = None
        if item.isdigit():
            dept = qs.filter(id=int(item)).first()
        if not dept:
            dept = qs.filter(code__iexact=item).first() or qs.filter(name__iexact=item).first()
        if dept:
            result.append(dept)
    return result


def _resolve_batches(value, request):
    if value is None or str(value).strip() == "":
        return []
    qs = Batch.objects.all()
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(department_id=request.user.department_id)
    items = [v.strip() for v in str(value).split(",") if v.strip()]
    result = []
    for item in items:
        batch = None
        if item.isdigit():
            batch = qs.filter(id=int(item)).first()
        if not batch:
            batch = qs.filter(name__iexact=item).first() or qs.filter(title__iexact=item).first()
        if not batch:
            for b in qs:
                if str(b).lower() == item.lower():
                    batch = b
                    break
        if batch:
            result.append(batch)
    return result


def _validate_session_rows(rows, request):
    cleaned = []
    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_header(k): v for k, v in row.items() }
        name = (normalized.get("name") or "").strip()
        start_date = _parse_date(normalized.get("start_date"))
        end_date = _parse_date(normalized.get("end_date"))
        is_active = _parse_bool(normalized.get("is_active"))
        departments = _resolve_departments(normalized.get("departments"), request)
        batches = _resolve_batches(normalized.get("batches"), request)

        if not name:
            errors.append("Name is required.")
        if not start_date:
            errors.append("Start Date is invalid or missing (use YYYY-MM-DD).")
        if not end_date:
            errors.append("End Date is invalid or missing (use YYYY-MM-DD).")

        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "name": name,
                "start_date": start_date.strftime("%Y-%m-%d") if start_date else "",
                "end_date": end_date.strftime("%Y-%m-%d") if end_date else "",
                "is_active": "true" if is_active else "false",
                "departments": ",".join([str(d.id) for d in departments]),
                "batches": ",".join([str(b.id) for b in batches]),
            },
            "clean": {
                "name": name,
                "start_date": start_date,
                "end_date": end_date,
                "is_active": is_active,
                "departments": departments,
                "batches": batches,
            },
        })
    return cleaned


def _create_session(clean):
    session = Session.objects.create(
        name=clean["name"],
        start_date=clean["start_date"],
        end_date=clean["end_date"],
        is_active=clean["is_active"],
    )
    if clean["departments"]:
        session.departments.set(clean["departments"])
    if clean["batches"]:
        session.batches.set(clean["batches"])


@login_required(login_url="accounts:login_page")
@never_cache
def sessions_template_download(request):
    deny = _require_sessions_read(request)
    if deny:
        return deny
    if not can(request.user, "SESSIONS", "create"):
        return redirect("academics:sessions")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sessions"
        ws.append(["Name", "Start Date", "End Date", "Is Active", "Departments", "Batches"])
        ws.append(["Fall 2024", "2024-09-01", "2024-12-31", "true", "CS", "BSCS 2022"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="sessions_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="sessions_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Name", "Start Date", "End Date", "Is Active", "Departments", "Batches"])
        writer.writerow(["Fall 2024", "2024-09-01", "2024-12-31", "true", "CS", "BSCS 2022"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def sessions_bulk_preview(request):
    deny = _require_sessions_read(request)
    if deny:
        return deny
    if not can(request.user, "SESSIONS", "create"):
        return redirect("academics:sessions")
    if request.method != "POST":
        return redirect("academics:sessions")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("academics:sessions")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("academics:sessions")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("academics:sessions")

    validated = _validate_session_rows(rows, request)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "name", "label": "Name", "type": "text"},
        {"key": "start_date", "label": "Start Date", "type": "date"},
        {"key": "end_date", "label": "End Date", "type": "date"},
        {"key": "is_active", "label": "Is Active", "type": "text"},
        {"key": "departments", "label": "Departments", "type": "text"},
        {"key": "batches", "label": "Batches", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "Session Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("academics:sessions_bulk_commit").url,
        "back_url": redirect("academics:sessions").url,
        "hidden_fields": [],
        "extra_badge": "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def sessions_bulk_commit(request):
    deny = _require_sessions_read(request)
    if deny:
        return deny
    if not can(request.user, "SESSIONS", "create"):
        return redirect("academics:sessions")
    if request.method != "POST":
        return redirect("academics:sessions")

    rows_json = request.POST.get("rows_json", "")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("academics:sessions")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("academics:sessions")

    validated = _validate_session_rows(rows, request)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "name", "label": "Name", "type": "text"},
            {"key": "start_date", "label": "Start Date", "type": "date"},
            {"key": "end_date", "label": "End Date", "type": "date"},
            {"key": "is_active", "label": "Is Active", "type": "text"},
            {"key": "departments", "label": "Departments", "type": "text"},
            {"key": "batches", "label": "Batches", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "Session Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("academics:sessions_bulk_commit").url,
            "back_url": redirect("academics:sessions").url,
            "hidden_fields": [],
            "extra_badge": "",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_session(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("academics:sessions")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_session(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("academics:sessions")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("academics:sessions")

    messages.error(request, "No valid rows to import.")
    return redirect("academics:sessions")


@login_required(login_url="accounts:login_page")
def create_session(request):
    deny = _require_sessions_read(request)
    if deny:
        return deny
    if request.method == "POST":
        if not can(request.user, "SESSIONS", "create"):
            return redirect("academics:sessions")
        form = SessionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("academics:sessions")
        qs = Session.objects.all().order_by("-start_date")
        paginator = Paginator(qs, 10)
        page_obj = paginator.get_page(None)
        return render(request, "academics/sessions.html", {
            "sessions": page_obj,
            "page_obj": page_obj,
            "form": form,
            "departments": Department.objects.filter(is_active=True).order_by("name"),
            "batches": Batch.objects.select_related("department").all().order_by("start_date"),
            "search": "",
        })
    return redirect("academics:sessions")


@login_required(login_url="accounts:login_page")
def edit_session(request, pk):
    deny = _require_sessions_read(request)
    if deny:
        return deny
    session = get_object_or_404(Session, pk=pk)
    if request.method == "POST":
        if not can(request.user, "SESSIONS", "update"):
            return redirect("academics:sessions")
        form = SessionForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
    return redirect("academics:sessions")


@login_required(login_url="accounts:login_page")
@require_POST
def delete_session(request, pk):
    deny = _require_sessions_read(request)
    if deny:
        return deny
    session = get_object_or_404(Session, pk=pk)
    if not can(request.user, "SESSIONS", "delete"):
        return redirect("academics:sessions")
    session.delete()
    return redirect("academics:sessions")
