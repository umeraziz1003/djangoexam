from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.contrib import messages
from django.db import IntegrityError, transaction
import json
import csv
from datetime import datetime, date
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from ..forms import BatchForm
from ..models import Batch, Department
from accounts.permissions import can


def _require_batches_read(request):
    if not can(request.user, "BATCHES", "read"):
        return redirect("accounts:dashboard")
    return None


@login_required(login_url="accounts:login_page")
@never_cache
def batches_view(request):
    deny = _require_batches_read(request)
    if deny:
        return deny
    search = request.GET.get("search", "").strip()
    department_id = request.GET.get("department_id", "")
    if request.user.is_department_scoped() and request.user.department_id:
        department_id = str(request.user.department_id)
    qs = Batch.objects.select_related("department").order_by("-start_date")
    if department_id:
        qs = qs.filter(department_id=department_id)
    if search:
        qs = qs.filter(
            Q(title__icontains=search)
            | Q(name__icontains=search)
            | Q(program__icontains=search)
        )

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    form = BatchForm()
    if request.user.is_department_scoped() and request.user.department_id:
        form.fields["department"].initial = request.user.department_id
        form.fields["department"].disabled = True
    departments = Department.objects.filter(is_active=True)
    if request.user.is_department_scoped() and request.user.department_id:
        departments = departments.filter(id=request.user.department_id)

    return render(request, "academics/batches.html", {
        "batches": page_obj,
        "page_obj": page_obj,
        "form": form,
        "search": search,
        "departments": departments,
        "department_id": department_id,
    })


_BATCH_HEADER_MAP = {
    "department": "department",
    "department_id": "department",
    "department_code": "department",
    "title": "title",
    "name": "name",
    "start_date": "start_date",
    "program": "program",
    "status": "status",
}


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _BATCH_HEADER_MAP.get(text, text)


def _parse_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    raw = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _resolve_department(value, request):
    if value is None or str(value).strip() == "":
        return None
    qs = Department.objects.all()
    if request.user.is_department_scoped() and request.user.department_id:
        qs = qs.filter(id=request.user.department_id)
    raw = str(value).strip()
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    direct = qs.filter(code__iexact=raw).first() or qs.filter(name__iexact=raw).first()
    return direct


def _validate_batch_rows(rows, request):
    cleaned = []
    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_header(k): v for k, v in row.items() }
        dept = _resolve_department(normalized.get("department"), request)
        title = (normalized.get("title") or "").strip()
        name = (normalized.get("name") or "").strip()
        program = (normalized.get("program") or "").strip()
        status = (normalized.get("status") or "ACTIVE").strip().upper()
        start_date = _parse_date(normalized.get("start_date"))

        if not dept:
            errors.append("Department is invalid.")
        if not title:
            errors.append("Title is required.")
        if not name:
            errors.append("Name is required.")
        if not program:
            errors.append("Program is required.")
        if not start_date:
            errors.append("Start Date is invalid or missing (use YYYY-MM-DD).")

        if status not in {"ACTIVE", "INACTIVE"}:
            status = "ACTIVE"

        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "department": str(dept.id) if dept else (normalized.get("department") or ""),
                "title": title,
                "name": name,
                "start_date": start_date.strftime("%Y-%m-%d") if start_date else "",
                "program": program,
                "status": status,
            },
            "clean": {
                "department": dept,
                "title": title,
                "name": name,
                "start_date": start_date,
                "program": program,
                "status": status,
            },
        })
    return cleaned


def _create_batch(clean):
    Batch.objects.create(
        department=clean["department"],
        title=clean["title"],
        name=clean["name"],
        start_date=clean["start_date"],
        program=clean["program"],
        status=clean["status"],
    )


@login_required(login_url="accounts:login_page")
@never_cache
def batches_template_download(request):
    deny = _require_batches_read(request)
    if deny:
        return deny
    if not can(request.user, "BATCHES", "create"):
        return redirect("academics:batches")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batches"
        ws.append(["Department", "Title", "Name", "Start Date", "Program", "Status"])
        ws.append(["CS", "Batch 2024", "Computer Science Batch 2024", "2024-09-01", "BSCS", "ACTIVE"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="batches_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="batches_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Department", "Title", "Name", "Start Date", "Program", "Status"])
        writer.writerow(["CS", "Batch 2024", "Computer Science Batch 2024", "2024-09-01", "BSCS", "ACTIVE"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def batches_bulk_preview(request):
    deny = _require_batches_read(request)
    if deny:
        return deny
    if not can(request.user, "BATCHES", "create"):
        return redirect("academics:batches")
    if request.method != "POST":
        return redirect("academics:batches")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("academics:batches")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("academics:batches")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("academics:batches")

    validated = _validate_batch_rows(rows, request)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "department", "label": "Department", "type": "text"},
        {"key": "title", "label": "Title", "type": "text"},
        {"key": "name", "label": "Name", "type": "text"},
        {"key": "start_date", "label": "Start Date", "type": "date"},
        {"key": "program", "label": "Program", "type": "text"},
        {"key": "status", "label": "Status", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "Batch Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("academics:batches_bulk_commit").url,
        "back_url": redirect("academics:batches").url,
        "hidden_fields": [],
        "extra_badge": "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def batches_bulk_commit(request):
    deny = _require_batches_read(request)
    if deny:
        return deny
    if not can(request.user, "BATCHES", "create"):
        return redirect("academics:batches")
    if request.method != "POST":
        return redirect("academics:batches")

    rows_json = request.POST.get("rows_json", "")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("academics:batches")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("academics:batches")

    validated = _validate_batch_rows(rows, request)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "department", "label": "Department", "type": "text"},
            {"key": "title", "label": "Title", "type": "text"},
            {"key": "name", "label": "Name", "type": "text"},
            {"key": "start_date", "label": "Start Date", "type": "date"},
            {"key": "program", "label": "Program", "type": "text"},
            {"key": "status", "label": "Status", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "Batch Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("academics:batches_bulk_commit").url,
            "back_url": redirect("academics:batches").url,
            "hidden_fields": [],
            "extra_badge": "",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_batch(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("academics:batches")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_batch(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("academics:batches")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("academics:batches")

    messages.error(request, "No valid rows to import.")
    return redirect("academics:batches")


@login_required(login_url="accounts:login_page")
def create_batch(request):
    deny = _require_batches_read(request)
    if deny:
        return deny
    if request.method == "POST":
        if not can(request.user, "BATCHES", "create"):
            return redirect("academics:batches")
        data = request.POST.copy()
        if request.user.is_department_scoped() and request.user.department_id:
            data["department"] = request.user.department_id
        form = BatchForm(data)
        if form.is_valid():
            if request.user.is_department_scoped() and request.user.department_id:
                if form.cleaned_data["department"].id != request.user.department_id:
                    return redirect("academics:batches")
            form.save()
            return redirect("academics:batches")
        qs = Batch.objects.select_related("department").order_by("-start_date")
        paginator = Paginator(qs, 10)
        page_obj = paginator.get_page(None)
        departments = Department.objects.filter(is_active=True)
        return render(request, "academics/batches.html", {
            "batches": page_obj,
            "page_obj": page_obj,
            "form": form,
            "search": "",
            "departments": departments,
        })
    return redirect("academics:batches")


@login_required(login_url="accounts:login_page")
def edit_batch(request, pk):
    deny = _require_batches_read(request)
    if deny:
        return deny
    batch = get_object_or_404(Batch, pk=pk)
    if request.method == "POST":
        if not can(request.user, "BATCHES", "update"):
            return redirect("academics:batches")
        if request.user.is_department_scoped() and request.user.department_id:
            if batch.department_id != request.user.department_id:
                return redirect("academics:batches")
        dept_id = request.POST.get("department_id")
        if dept_id and not (request.user.is_department_scoped() and request.user.department_id):
            batch.department_id = dept_id
        batch.title = request.POST.get("title", batch.title).strip()
        batch.name = request.POST.get("name", batch.name).strip()
        batch.start_date = request.POST.get("start_date", batch.start_date)
        batch.program = request.POST.get("program", batch.program).strip()
        batch.status = request.POST.get("status", batch.status)
        batch.save()
    return redirect("academics:batches")


@login_required(login_url="accounts:login_page")
@require_POST
def delete_batch(request, pk):
    deny = _require_batches_read(request)
    if deny:
        return deny
    batch = get_object_or_404(Batch, pk=pk)
    if not can(request.user, "BATCHES", "delete"):
        return redirect("academics:batches")
    if request.user.is_department_scoped() and request.user.department_id:
        if batch.department_id != request.user.department_id:
            return redirect("academics:batches")
    batch.delete()
    return redirect("academics:batches")
