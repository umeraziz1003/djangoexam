from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import Group, Permission
from django.shortcuts import redirect, render, get_object_or_404
from django.views.decorators.cache import never_cache
from django.http import HttpResponse
from django.db import IntegrityError, transaction
from django.contrib import messages
import json
import csv

from admission.models import Student
from .permissions import ensure_default_groups, reset_default_permissions, MODULE_PERMS, MODULE_LABELS, ACTIONS, perm_codename, DEFAULT_GROUPS
from django.contrib.auth import get_user_model
from academics.models import Department

@never_cache
def logout_view(request):
    if request.method == "POST":
        logout(request)
        return redirect("accounts:login_page")
    return render(request, "accounts/logout.html")


def login_page(request):
    if request.user.is_authenticated:
        return redirect("accounts:dashboard")

    error = None
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)
        if user is not None and not getattr(user, "is_locked", False):
            login(request, user)
            return redirect("accounts:dashboard")
        if user is not None and getattr(user, "is_locked", False):
            error = "Your account is locked. Please contact the administrator."
        else:
            error = "Invalid username or password. Please try again."

    return render(request, "accounts/login.html", {"error": error})


@login_required(login_url="accounts:login_page")
@never_cache
def dashboard(request):
    from academics.models import Department, Batch, Session
    from results.models import Result

    student_profile = None
    student_results = []
    student_cgpa = None
    dept_count = batch_count = session_count = student_count = 0

    if request.user.is_student():
        try:
            student_profile = request.user.student_profile
        except Student.DoesNotExist:
            student_profile = None
        if student_profile:
            student_results = Result.objects.select_related(
                "enrollment",
                "enrollment__course_offering",
                "enrollment__course_offering__course",
                "enrollment__course_offering__session",
            ).filter(
                enrollment__student=student_profile,
                result_published=True,
            ).order_by("-calculated_at")[:6]

    if request.user.is_exam_officer() or request.user.is_dept_controller():
        dept_count = Department.objects.filter(is_active=True).count()
        batch_count = Batch.objects.filter(status="ACTIVE").count()
        session_count = Session.objects.filter(is_active=True).count()
        student_count = Student.objects.filter(is_active=True).count()

    is_student = request.user.is_student()
    is_exam_officer = request.user.is_exam_officer()

    context = {
        "student_profile": student_profile,
        "student_results": student_results,
        "student_cgpa": student_cgpa,
        "is_student": is_student,
        "is_exam_officer": is_exam_officer,
        "dept_count": dept_count,
        "batch_count": batch_count,
        "session_count": session_count,
        "student_count": student_count,
    }
    return render(request, "dashboard/index.html", context)


@login_required(login_url="accounts:login_page")
@never_cache
def permissions_view(request):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")

    ensure_default_groups()
    groups = Group.objects.all().order_by("name")

    if request.method == "POST":
        if request.POST.get("action") == "reset":
            reset_default_permissions()
            return redirect("accounts:permissions")

        for group in groups:
            perm_objs = []
            for module in MODULE_PERMS.keys():
                key = f"{group.id}__{module}"
                for action in ACTIONS:
                    flag = key + f"__{action[0]}"
                    if flag in request.POST:
                        code = perm_codename(module, action)
                        if code:
                            app_label, codename = code
                            perm = Permission.objects.filter(
                                content_type__app_label=app_label,
                                codename=codename,
                            ).first()
                            if perm:
                                perm_objs.append(perm)
            group.permissions.set(perm_objs)
        return redirect("accounts:permissions")

    rows = []
    for group in groups:
        for module in MODULE_PERMS.keys():
            row = {
                "group": group,
                "module": module,
                "module_label": MODULE_LABELS.get(module, module),
            }
            for action in ACTIONS:
                code = perm_codename(module, action)
                if not code:
                    row[f"can_{action}"] = False
                    continue
                app_label, codename = code
                row[f"can_{action}"] = group.permissions.filter(
                    content_type__app_label=app_label,
                    codename=codename,
                ).exists()
            rows.append(row)

    return render(request, "accounts/permissions.html", {"perms": rows})


@login_required(login_url="accounts:login_page")
@never_cache
def groups_view(request):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    ensure_default_groups()

    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "add":
            name = (request.POST.get("name", "") or "").strip()
            if not name:
                messages.error(request, "Group name is required.")
            elif Group.objects.filter(name=name).exists():
                messages.error(request, "Group already exists.")
            else:
                Group.objects.create(name=name)
                messages.success(request, "Group created.")
            return redirect("accounts:groups")
        if action == "edit":
            group_id = request.POST.get("group_id")
            group = get_object_or_404(Group, pk=group_id)
            if group.name in DEFAULT_GROUPS:
                messages.error(request, "System groups cannot be edited.")
                return redirect("accounts:groups")
            name = (request.POST.get("name", group.name) or "").strip()
            if not name:
                messages.error(request, "Group name is required.")
                return redirect("accounts:groups")
            group.name = name
            group.save(update_fields=["name"])
            messages.success(request, "Group updated.")
            return redirect("accounts:groups")
        if action == "delete":
            group_id = request.POST.get("group_id")
            group = get_object_or_404(Group, pk=group_id)
            if group.name in DEFAULT_GROUPS:
                messages.error(request, "System groups cannot be deleted.")
                return redirect("accounts:groups")
            if group.user_set.exists():
                messages.error(request, "Group is assigned to users and cannot be deleted.")
                return redirect("accounts:groups")
            group.delete()
            messages.success(request, "Group deleted.")
            return redirect("accounts:groups")

    groups = Group.objects.all().order_by("name")
    return render(request, "accounts/groups.html", {
        "groups": groups,
        "system_group_names": list(DEFAULT_GROUPS.keys()),
    })


@login_required(login_url="accounts:login_page")
@never_cache
def users_view(request):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")

    User = get_user_model()
    search = request.GET.get("search", "").strip()
    group_id = request.GET.get("group", "")
    department_id = request.GET.get("department_id", "")

    qs = User.objects.select_related("department").order_by("username")
    if group_id:
        qs = qs.filter(groups__id=group_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if search:
        qs = qs.filter(username__icontains=search)

    return render(request, "accounts/users.html", {
        "users": qs,
        "search": search,
        "group": group_id,
        "department_id": department_id,
        "departments": Department.objects.filter(is_active=True).order_by("name"),
        "groups": Group.objects.all().order_by("name"),
    })


@login_required(login_url="accounts:login_page")
@never_cache
def create_user(request):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")

    User = get_user_model()
    error = None
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        group_ids = request.POST.getlist("groups")
        department_id = request.POST.get("department_id") or None

        if not username or not password or not group_ids:
            error = "Username, password, and group are required."
        elif User.objects.filter(username=username).exists():
            error = "A user with this username already exists."
        else:
            valid_groups = Group.objects.filter(id__in=group_ids)
            if not valid_groups.exists():
                error = "Selected groups are not valid."
                return render(request, "accounts/create_user.html", {
                    "error": error,
                    "groups": Group.objects.all().order_by("name"),
                    "departments": Department.objects.filter(is_active=True).order_by("name"),
                })
            user = User.objects.create_user(
                username=username,
                password=password,
                department_id=department_id,
            )
            user.groups.set(valid_groups)
            return redirect("accounts:users")

    return render(request, "accounts/create_user.html", {
        "error": error,
        "groups": Group.objects.all().order_by("name"),
        "departments": Department.objects.filter(is_active=True).order_by("name"),
    })


@login_required(login_url="accounts:login_page")
@never_cache
def edit_user(request, pk):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")

    User = get_user_model()
    user = get_object_or_404(User, pk=pk)
    error = None
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        group_ids = request.POST.getlist("groups")
        department_id = request.POST.get("department_id") or None
        password = request.POST.get("password", "")

        if not username:
            error = "Username is required."
        else:
            if User.objects.filter(username=username).exclude(pk=user.pk).exists():
                error = "A user with this username already exists."
            else:
                user.username = username
                if group_ids:
                    user.groups.set(Group.objects.filter(id__in=group_ids))
                user.department_id = department_id
                if password:
                    user.set_password(password)
                user.save()
                return redirect("accounts:users")

    groups_qs = Group.objects.all().order_by("name")
    return render(request, "accounts/edit_user.html", {
        "user_obj": user,
        "error": error,
        "groups": groups_qs,
        "departments": Department.objects.filter(is_active=True).order_by("name"),
    })


@login_required(login_url="accounts:login_page")
@never_cache
def delete_user(request, pk):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")

    User = get_user_model()
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        if user.pk == request.user.pk:
            return redirect("accounts:users")
        user.delete()
        return redirect("accounts:users")
    return render(request, "accounts/delete_user.html", {"user_obj": user})


@login_required(login_url="accounts:login_page")
@never_cache
def toggle_user_lock(request, pk):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    User = get_user_model()
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        if user.pk == request.user.pk:
            return redirect("accounts:users")
        user.is_locked = not user.is_locked
        user.save(update_fields=["is_locked"])
    return redirect("accounts:users")


@login_required(login_url="accounts:login_page")
@never_cache
def reset_user_password(request, pk):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    User = get_user_model()
    user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        new_password = request.POST.get("password", "")
        if new_password:
            user.set_password(new_password)
            user.save(update_fields=["password"])
        return redirect("accounts:users")
    return render(request, "accounts/reset_user_password.html", {"user_obj": user})


def sample_view(request):
    from django.http import HttpResponse
    return HttpResponse("accounts sample view")


_USER_HEADER_MAP = {
    "username": "username",
    "user": "username",
    "password": "password",
    "groups": "groups",
    "group": "groups",
    "department": "department",
    "department_id": "department",
}


def _normalize_header(header):
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace(" ", "_")
    return _USER_HEADER_MAP.get(text, text)


def _resolve_groups(value):
    if value is None or str(value).strip() == "":
        return []
    items = [v.strip() for v in str(value).split(",") if v.strip()]
    groups = []
    for item in items:
        group = None
        if item.isdigit():
            group = Group.objects.filter(id=int(item)).first()
        if not group:
            group = Group.objects.filter(name__iexact=item).first()
        if group:
            groups.append(group)
    return groups


def _resolve_department(value):
    if value is None or str(value).strip() == "":
        return None
    raw = str(value).strip()
    qs = Department.objects.filter(is_active=True)
    if raw.isdigit():
        return qs.filter(id=int(raw)).first()
    return qs.filter(name__iexact=raw).first() or qs.filter(code__iexact=raw).first()


def _validate_user_rows(rows):
    cleaned = []
    User = get_user_model()
    usernames = [str(r.get("username") or "").strip() for r in rows if r.get("username")]
    existing = set(User.objects.filter(username__in=usernames).values_list("username", flat=True))
    seen = set()

    for idx, row in enumerate(rows, start=1):
        errors = []
        normalized = { _normalize_header(k): v for k, v in row.items() }
        username = (normalized.get("username") or "").strip()
        password = (normalized.get("password") or "").strip()
        groups = _resolve_groups(normalized.get("groups"))
        department = _resolve_department(normalized.get("department"))

        if not username:
            errors.append("Username is required.")
        if not password:
            errors.append("Password is required.")
        if not groups:
            errors.append("Groups are required.")
        if username in seen:
            errors.append("Username is duplicated in the upload.")
        if username in existing:
            errors.append("Username already exists.")

        seen.add(username)

        cleaned.append({
            "row_num": idx,
            "errors": errors,
            "raw": normalized,
            "display": {
                "username": username,
                "password": password,
                "groups": ",".join([g.name for g in groups]),
                "department": str(department.id) if department else "",
            },
            "clean": {
                "username": username,
                "password": password,
                "groups": groups,
                "department": department,
            },
        })
    return cleaned


def _create_user(clean):
    User = get_user_model()
    user = User.objects.create_user(
        username=clean["username"],
        password=clean["password"],
        department=clean["department"],
    )
    if clean["groups"]:
        user.groups.set(clean["groups"])


@login_required(login_url="accounts:login_page")
@never_cache
def users_template_download(request):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Username", "Password", "Groups", "Department"])
        ws.append(["dept_controller2", "pass123", "DEPT_CONTROLLER", "CS"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="users_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="users_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Username", "Password", "Groups", "Department"])
        writer.writerow(["dept_controller2", "pass123", "DEPT_CONTROLLER", "CS"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def users_bulk_preview(request):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    if request.method != "POST":
        return redirect("accounts:create_user")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("accounts:create_user")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("accounts:create_user")
        header_keys = [_normalize_header(h) for h in headers]
        rows = []
        for row in rows_iter:
            if row is None:
                continue
            row_dict = {}
            empty = True
            for i, key in enumerate(header_keys):
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                row_dict[key] = val
            if empty:
                continue
            rows.append(row_dict)
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("accounts:create_user")

    validated = _validate_user_rows(rows)
    total = len(validated)
    invalid = sum(1 for r in validated if r["errors"])
    valid = total - invalid

    columns = [
        {"key": "username", "label": "Username", "type": "text"},
        {"key": "password", "label": "Password", "type": "text"},
        {"key": "groups", "label": "Groups", "type": "text"},
        {"key": "department", "label": "Department", "type": "text"},
    ]
    return render(request, "shared/bulk_preview.html", {
        "page_title": "User Import Preview",
        "rows": validated,
        "columns": columns,
        "total": total,
        "valid_count": valid,
        "invalid_count": invalid,
        "rows_json": json.dumps(rows, default=str),
        "source": "excel",
        "commit_url": redirect("accounts:users_bulk_commit").url,
        "back_url": redirect("accounts:create_user").url,
        "hidden_fields": [],
        "extra_badge": "",
    })


@login_required(login_url="accounts:login_page")
@never_cache
def users_bulk_commit(request):
    if not request.user.is_exam_officer():
        return redirect("accounts:dashboard")
    if request.method != "POST":
        return redirect("accounts:create_user")

    rows_json = request.POST.get("rows_json", "")
    import_valid = request.POST.get("import_valid") == "1"

    try:
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError:
        messages.error(request, "Bulk data is corrupted. Please upload again.")
        return redirect("accounts:create_user")

    if not rows:
        messages.error(request, "No rows to import.")
        return redirect("accounts:create_user")

    validated = _validate_user_rows(rows)
    invalid_rows = [r for r in validated if r["errors"]]
    valid_rows = [r for r in validated if not r["errors"]]

    if invalid_rows and not import_valid:
        total = len(validated)
        invalid = len(invalid_rows)
        valid = total - invalid
        messages.error(request, "Import blocked due to validation errors. Fix rows or import only valid rows.")
        columns = [
            {"key": "username", "label": "Username", "type": "text"},
            {"key": "password", "label": "Password", "type": "text"},
            {"key": "groups", "label": "Groups", "type": "text"},
            {"key": "department", "label": "Department", "type": "text"},
        ]
        return render(request, "shared/bulk_preview.html", {
            "page_title": "User Import Preview",
            "rows": validated,
            "columns": columns,
            "total": total,
            "valid_count": valid,
            "invalid_count": invalid,
            "rows_json": json.dumps(rows, default=str),
            "source": "excel",
            "commit_url": redirect("accounts:users_bulk_commit").url,
            "back_url": redirect("accounts:create_user").url,
            "hidden_fields": [],
            "extra_badge": "",
        })

    created = 0
    failed = 0
    if invalid_rows and import_valid:
        for r in valid_rows:
            try:
                _create_user(r["clean"])
                created += 1
            except IntegrityError:
                failed += 1
        if failed:
            messages.warning(request, f"Imported {created} rows. {failed} rows failed during save.")
        else:
            messages.success(request, f"Imported {created} rows successfully.")
        return redirect("accounts:users")

    if valid_rows and not invalid_rows:
        try:
            with transaction.atomic():
                for r in valid_rows:
                    _create_user(r["clean"])
                    created += 1
            messages.success(request, f"Imported {created} rows successfully.")
            return redirect("accounts:users")
        except IntegrityError:
            messages.error(request, "Import failed due to a database conflict. No rows were saved.")
            return redirect("accounts:create_user")

    messages.error(request, "No valid rows to import.")
    return redirect("accounts:create_user")
