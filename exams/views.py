from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.cache import never_cache
from django.http import HttpResponse
from django.db import transaction
from django.contrib import messages
import csv

from enrollments.models import Enrollment
from courses.models import CourseOffering
from results.models import Result

from .forms import ExamSplitConfigForm, GradeScaleForm, ExamRulesForm
from .models import ExamSplitConfig, GradeScale, ExamRules, Marks, EXAM_TYPE_CHOICES
from accounts.permissions import can


def _get_grade_for_total(total):
    scale = GradeScale.objects.filter(is_active=True, min_percentage__lte=total, max_percentage__gte=total) \
        .order_by("-min_percentage") \
        .first()
    if not scale:
        return "N/A", 0
    return scale.grade, scale.grade_point


def _require_exam_read(request):
    if not can(request.user, "EXAMS", "read"):
        return redirect("accounts:dashboard")
    return None


def _parse_float(value, label, row, errors, min_val=None, max_val=None):
    if value is None or str(value).strip() == "":
        errors.append(f"Row {row}: {label} is required.")
        return None
    try:
        num = float(value)
    except ValueError:
        errors.append(f"Row {row}: {label} must be a number.")
        return None
    if min_val is not None and num < min_val:
        errors.append(f"Row {row}: {label} must be at least {min_val}.")
        return None
    if max_val is not None and num > max_val:
        errors.append(f"Row {row}: {label} must be at most {max_val}.")
        return None
    return num


def _parse_bool(value):
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "on"}


def _save_grade_scale_rows(rows):
    created = 0
    for row in rows:
        GradeScale.objects.create(
            min_percentage=row["min_percentage"],
            max_percentage=row["max_percentage"],
            grade=row["grade"],
            grade_point=row["grade_point"],
            is_active=row["is_active"],
        )
        created += 1
    return created


@login_required(login_url="accounts:login_page")
@never_cache
def marks_template_download(request):
    deny = _require_exam_read(request)
    if deny:
        return deny
    if not can(request.user, "EXAMS", "update"):
        return redirect("exams:manage_marks")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marks"
        ws.append(["Roll No", "Sessional", "Midterm", "Terminal"])
        ws.append(["CS-2024-001", "12", "18", "40"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="marks_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="marks_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Roll No", "Sessional", "Midterm", "Terminal"])
        writer.writerow(["CS-2024-001", "12", "18", "40"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def exam_splits_view(request):
    deny = _require_exam_read(request)
    if deny:
        return deny
    config = ExamSplitConfig.get_solo()
    if request.method == "POST":
        form = ExamSplitConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam splits updated.")
            return redirect("exams:exam_splits")
    else:
        form = ExamSplitConfigForm(instance=config)
    return render(request, "exams/exam_splits.html", {"form": form, "total": config.total()})


@login_required(login_url="accounts:login_page")
@never_cache
def grading_system_view(request):
    deny = _require_exam_read(request)
    if deny:
        return deny
    if request.method == "POST":
        action = request.POST.get("action") or "add_one"
        if action == "bulk_add":
            if not can(request.user, "EXAMS", "update"):
                return redirect("exams:grading_system")
            row_count_raw = request.POST.get("row_count") or "0"
            try:
                row_count = int(row_count_raw)
            except ValueError:
                row_count = 0
            rows = []
            errors = []
            for i in range(1, row_count + 1):
                row_errors = []
                min_val = request.POST.get(f"min_percentage_{i}")
                max_val = request.POST.get(f"max_percentage_{i}")
                grade = (request.POST.get(f"grade_{i}") or "").strip()
                grade_point = request.POST.get(f"grade_point_{i}")
                is_active = request.POST.get(f"is_active_{i}")

                if not any([min_val, max_val, grade, grade_point, is_active]):
                    continue

                min_num = _parse_float(min_val, "Min %", i, row_errors, min_val=0, max_val=100)
                max_num = _parse_float(max_val, "Max %", i, row_errors, min_val=0, max_val=100)
                gp_num = _parse_float(grade_point, "Grade Point", i, row_errors, min_val=0, max_val=4)

                if grade == "":
                    row_errors.append(f"Row {i}: Grade is required.")

                if min_num is not None and max_num is not None and min_num > max_num:
                    row_errors.append(f"Row {i}: Min % cannot be greater than Max %.")

                if row_errors:
                    errors.extend(row_errors)
                    continue

                rows.append({
                    "min_percentage": min_num,
                    "max_percentage": max_num,
                    "grade": grade,
                    "grade_point": gp_num,
                    "is_active": _parse_bool(is_active),
                })

            if not rows and not errors:
                messages.error(request, "No rows to add.")
                return redirect("exams:grading_system")
            if errors:
                for msg in errors[:10]:
                    messages.error(request, msg)
                if len(errors) > 10:
                    messages.error(request, f"{len(errors) - 10} more errors not shown.")
                return redirect("exams:grading_system")

            with transaction.atomic():
                created = _save_grade_scale_rows(rows)
            messages.success(request, f"Added {created} grade scale rows.")
            return redirect("exams:grading_system")

        form = GradeScaleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Grade scale added.")
            return redirect("exams:grading_system")
    else:
        form = GradeScaleForm()

    scales = GradeScale.objects.all()
    return render(request, "exams/grading_system.html", {"form": form, "scales": scales})


@login_required(login_url="accounts:login_page")
@never_cache
def grade_scale_template_download(request):
    deny = _require_exam_read(request)
    if deny:
        return deny
    if not can(request.user, "EXAMS", "update"):
        return redirect("exams:grading_system")
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grade Scales"
        ws.append(["Min %", "Max %", "Grade", "Grade Point", "Active"])
        ws.append(["85", "100", "A", "4.00", "Yes"])
        ws.append(["80", "84.99", "A-", "3.67", "Yes"])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="grade_scale_template.xlsx"'
        return response
    except Exception:
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="grade_scale_template.csv"'
        writer = csv.writer(response)
        writer.writerow(["Min %", "Max %", "Grade", "Grade Point", "Active"])
        writer.writerow(["85", "100", "A", "4.00", "Yes"])
        writer.writerow(["80", "84.99", "A-", "3.67", "Yes"])
        return response


@login_required(login_url="accounts:login_page")
@never_cache
def grade_scale_import(request):
    deny = _require_exam_read(request)
    if deny:
        return deny
    if not can(request.user, "EXAMS", "update"):
        return redirect("exams:grading_system")
    if request.method != "POST":
        return redirect("exams:grading_system")

    upload = request.FILES.get("excel_file")
    if not upload:
        messages.error(request, "Please choose an Excel file to upload.")
        return redirect("exams:grading_system")

    try:
        import openpyxl
    except ImportError:
        messages.error(request, "Excel import requires openpyxl. Please install it first.")
        return redirect("exams:grading_system")

    errors = []
    rows = []
    try:
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            messages.error(request, "Excel file is empty.")
            return redirect("exams:grading_system")
        header_map = {}
        for idx, h in enumerate(headers):
            key = str(h).strip().lower().replace(" ", "_").replace("%", "")
            header_map[key] = idx

        for row_idx, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            row_errors = []
            def _get_val(keys):
                for k in keys:
                    if k in header_map:
                        return row[header_map[k]]
                return None

            min_val = _get_val(["min", "min_", "min_percentage", "minpercent", "minpercentage"])
            max_val = _get_val(["max", "max_", "max_percentage", "maxpercent", "maxpercentage"])
            grade = _get_val(["grade"])
            grade_point = _get_val(["grade_point", "gradepoint", "points", "point"])
            active = _get_val(["active", "is_active", "status"])

            if all(v in (None, "") for v in [min_val, max_val, grade, grade_point, active]):
                continue

            min_num = _parse_float(min_val, "Min %", row_idx, row_errors, min_val=0, max_val=100)
            max_num = _parse_float(max_val, "Max %", row_idx, row_errors, min_val=0, max_val=100)
            gp_num = _parse_float(grade_point, "Grade Point", row_idx, row_errors, min_val=0, max_val=4)
            grade_text = (str(grade).strip() if grade is not None else "")

            if grade_text == "":
                row_errors.append(f"Row {row_idx}: Grade is required.")
            if min_num is not None and max_num is not None and min_num > max_num:
                row_errors.append(f"Row {row_idx}: Min % cannot be greater than Max %.")

            if row_errors:
                errors.extend(row_errors)
                continue

            rows.append({
                "min_percentage": min_num,
                "max_percentage": max_num,
                "grade": grade_text,
                "grade_point": gp_num,
                "is_active": _parse_bool(active),
            })
    except Exception:
        messages.error(request, "Failed to read the Excel file. Please check its format.")
        return redirect("exams:grading_system")

    if errors:
        for msg in errors[:10]:
            messages.error(request, msg)
        if len(errors) > 10:
            messages.error(request, f"{len(errors) - 10} more errors not shown.")
        return redirect("exams:grading_system")

    if not rows:
        messages.error(request, "No valid rows found in the Excel file.")
        return redirect("exams:grading_system")

    with transaction.atomic():
        created = _save_grade_scale_rows(rows)
    messages.success(request, f"Imported {created} grade scale rows.")
    return redirect("exams:grading_system")


@login_required(login_url="accounts:login_page")
@never_cache
def edit_grade_scale(request, pk):
    deny = _require_exam_read(request)
    if deny:
        return deny
    scale = get_object_or_404(GradeScale, pk=pk)
    if request.method == "POST":
        form = GradeScaleForm(request.POST, instance=scale)
        if form.is_valid():
            form.save()
            messages.success(request, "Grade scale updated.")
            return redirect("exams:grading_system")
    else:
        form = GradeScaleForm(instance=scale)
    return render(request, "exams/edit_grade_scale.html", {"form": form, "scale": scale})


@login_required(login_url="accounts:login_page")
@never_cache
def delete_grade_scale(request, pk):
    deny = _require_exam_read(request)
    if deny:
        return deny
    scale = get_object_or_404(GradeScale, pk=pk)
    if request.method == "POST":
        scale.delete()
        messages.success(request, "Grade scale removed.")
        return redirect("exams:grading_system")
    return render(request, "exams/delete_grade_scale.html", {"scale": scale})


@login_required(login_url="accounts:login_page")
@never_cache
def exam_rules_view(request):
    deny = _require_exam_read(request)
    if deny:
        return deny
    rules = ExamRules.get_solo()
    if request.method == "POST":
        form = ExamRulesForm(request.POST, instance=rules)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam rules updated.")
            return redirect("exams:exam_rules")
    else:
        form = ExamRulesForm(instance=rules)
    return render(request, "exams/exam_rules.html", {"form": form})


@login_required(login_url="accounts:login_page")
@never_cache
def manage_marks_view(request):
    deny = _require_exam_read(request)
    if deny:
        return deny
    config = ExamSplitConfig.get_solo()
    offerings = CourseOffering.objects.select_related("course", "session", "semester").order_by("course__course_code")
    if request.user.is_department_scoped() and request.user.department_id:
        offerings = offerings.filter(course__department_id=request.user.department_id)

    offering_id = request.GET.get("offering_id") or request.POST.get("offering_id") or ""
    enrollments = Enrollment.objects.select_related("student", "course_offering").filter(
        course_offering_id=offering_id
    ).order_by("student__roll_no") if offering_id else Enrollment.objects.none()
    if request.user.is_department_scoped() and request.user.department_id:
        enrollments = enrollments.filter(course_offering__course__department_id=request.user.department_id)

    marks_map = {(m.enrollment_id, m.exam_type): m for m in Marks.objects.filter(enrollment__in=enrollments)}
    rows = []
    for enrollment in enrollments:
        rows.append({
            "enrollment": enrollment,
            "sessional": marks_map.get((enrollment.id, "SESSIONAL")).obtained_marks if marks_map.get((enrollment.id, "SESSIONAL")) else "",
            "midterm": marks_map.get((enrollment.id, "MIDTERM")).obtained_marks if marks_map.get((enrollment.id, "MIDTERM")) else "",
            "terminal": marks_map.get((enrollment.id, "TERMINAL")).obtained_marks if marks_map.get((enrollment.id, "TERMINAL")) else "",
        })

    if request.method == "POST" and request.POST.get("action") == "upload_excel":
        if not can(request.user, "EXAMS", "update"):
            return redirect("exams:manage_marks")
        upload = request.FILES.get("excel_file")
        if not offering_id or not upload:
            messages.error(request, "Please select an offering and choose an Excel file.")
            return redirect("exams:manage_marks")

        try:
            import openpyxl
        except ImportError:
            messages.error(request, "Excel import requires openpyxl. Please install it first.")
            return redirect("exams:manage_marks")

        enrollments = Enrollment.objects.select_related("student", "course_offering").filter(
            course_offering_id=offering_id
        ).order_by("student__roll_no")
        if request.user.is_department_scoped() and request.user.department_id:
            enrollments = enrollments.filter(course_offering__course__department_id=request.user.department_id)

        enrollment_by_roll = {e.student.roll_no: e for e in enrollments}
        errors = []
        updates = []

        try:
            wb = openpyxl.load_workbook(upload, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if not headers:
                messages.error(request, "Excel file is empty.")
                return redirect("exams:manage_marks")
            header_map = {}
            for idx, h in enumerate(headers):
                key = str(h).strip().lower().replace(" ", "_")
                header_map[key] = idx

            for row_idx, row in enumerate(rows_iter, start=2):
                if row is None:
                    continue
                roll_no = None
                if "roll_no" in header_map:
                    roll_no = row[header_map["roll_no"]]
                elif "roll" in header_map:
                    roll_no = row[header_map["roll"]]
                if not roll_no:
                    continue
                roll_no = str(roll_no).strip()
                enrollment = enrollment_by_roll.get(roll_no)
                if not enrollment:
                    errors.append(f"Row {row_idx}: Roll No {roll_no} not found in this offering.")
                    continue

                def _get_val(keys):
                    for k in keys:
                        if k in header_map:
                            return row[header_map[k]]
                    return None

                sessional = _get_val(["sessional"])
                midterm = _get_val(["midterm"])
                terminal = _get_val(["terminal"])

                def _parse_mark(val, max_allowed, label):
                    if val is None or str(val).strip() == "":
                        return None
                    try:
                        num = float(val)
                    except ValueError:
                        errors.append(f"Row {row_idx}: Invalid {label} marks for {roll_no}.")
                        return None
                    if num < 0 or num > max_allowed:
                        errors.append(f"Row {row_idx}: {label} marks must be 0 - {max_allowed} for {roll_no}.")
                        return None
                    return num

                sessional_val = _parse_mark(sessional, config.sessional_max, "Sessional")
                midterm_val = _parse_mark(midterm, config.midterm_max, "Midterm")
                terminal_val = _parse_mark(terminal, config.terminal_max, "Terminal")

                updates.append((enrollment, sessional_val, midterm_val, terminal_val))
        except Exception:
            messages.error(request, "Failed to read the Excel file. Please check its format.")
            return redirect("exams:manage_marks")

        if errors:
            for msg in errors[:10]:
                messages.error(request, msg)
            if len(errors) > 10:
                messages.error(request, f"{len(errors) - 10} more errors not shown.")
            return redirect(f"{request.path}?offering_id={offering_id}")

        with transaction.atomic():
            for enrollment, sessional_val, midterm_val, terminal_val in updates:
                for exam_type, value in [("SESSIONAL", sessional_val), ("MIDTERM", midterm_val), ("TERMINAL", terminal_val)]:
                    if value is None:
                        continue
                    Marks.objects.update_or_create(
                        enrollment=enrollment,
                        exam_type=exam_type,
                        defaults={
                            "obtained_marks": value,
                            "entered_by": request.user,
                        },
                    )

                total = 0
                for exam_type, _ in EXAM_TYPE_CHOICES:
                    mark = Marks.objects.filter(enrollment=enrollment, exam_type=exam_type).first()
                    if mark:
                        total += mark.obtained_marks
                grade, grade_point = _get_grade_for_total(total)
                result, created = Result.objects.get_or_create(
                    enrollment=enrollment,
                    defaults={
                        "total_marks": total,
                        "grade": grade,
                        "grade_point": grade_point,
                        "result_published": False,
                    }
                )
                if not created:
                    result.total_marks = total
                    result.grade = grade
                    result.grade_point = grade_point
                    result.save(update_fields=["total_marks", "grade", "grade_point", "calculated_at"])

        messages.success(request, "Marks imported successfully.")
        return redirect(f"{request.path}?offering_id={offering_id}")

    if request.method == "POST" and offering_id:
        if not can(request.user, "EXAMS", "update"):
            return redirect("exams:manage_marks")
        errors = []
        for enrollment in enrollments:
            for exam_type, label in EXAM_TYPE_CHOICES:
                key = f"{exam_type.lower()}_{enrollment.id}"
                raw = request.POST.get(key, "").strip()
                if raw == "":
                    continue
                try:
                    value = float(raw)
                except ValueError:
                    errors.append(f"Invalid marks for {enrollment.student.roll_no} ({label}).")
                    continue

                max_allowed = {
                    "SESSIONAL": config.sessional_max,
                    "MIDTERM": config.midterm_max,
                    "TERMINAL": config.terminal_max,
                }[exam_type]

                if value < 0 or value > max_allowed:
                    errors.append(f"{enrollment.student.roll_no} {label} must be 0 - {max_allowed}.")
                    continue

                Marks.objects.update_or_create(
                    enrollment=enrollment,
                    exam_type=exam_type,
                    defaults={
                        "obtained_marks": value,
                        "entered_by": request.user,
                    },
                )

            # calculate final result
            total = 0
            for exam_type, _ in EXAM_TYPE_CHOICES:
                mark = Marks.objects.filter(enrollment=enrollment, exam_type=exam_type).first()
                if mark:
                    total += mark.obtained_marks

            grade, grade_point = _get_grade_for_total(total)
            result, created = Result.objects.get_or_create(
                enrollment=enrollment,
                defaults={
                    "total_marks": total,
                    "grade": grade,
                    "grade_point": grade_point,
                    "result_published": False,
                }
            )
            if not created:
                result.total_marks = total
                result.grade = grade
                result.grade_point = grade_point
                result.save(update_fields=["total_marks", "grade", "grade_point", "calculated_at"])

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            messages.success(request, "Marks saved and results updated.")
        return redirect(f"{request.path}?offering_id={offering_id}")

    context = {
        "offerings": offerings,
        "offering_id": offering_id,
        "enrollments": enrollments,
        "rows": rows,
        "splits": config,
    }
    return render(request, "exams/manage_marks.html", context)
